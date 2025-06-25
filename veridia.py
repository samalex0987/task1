
#!/usr/bin/env python   
# -*- coding: utf-8 -*-
import streamlit as st
import requests
from veridia1 import *
import json 
import urllib.parse
import urllib3
import certifi
import pandas as pd
from datetime import datetime
import re
import logging
import os
from dotenv import load_dotenv
import aiohttp
import asyncio
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import openpyxl
import io
from uuid import uuid4
import ibm_boto3
from ibm_botocore.client import Config
from tenacity import retry, stop_after_attempt, wait_exponential
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
import traceback
from veridia1 import *
from datetime import date
import concurrent.futures
from dateutil.relativedelta import relativedelta

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Load environment variables
load_dotenv()

# IBM COS Configuration
COS_API_KEY = os.getenv("COS_API_KEY")
COS_SERVICE_INSTANCE_ID = os.getenv("COS_SERVICE_INSTANCE_ID")
COS_ENDPOINT = os.getenv("COS_ENDPOINT")
COS_BUCKET = os.getenv("COS_BUCKET")

# WatsonX configuration
WATSONX_API_URL = os.getenv("WATSONX_API_URL_1")
MODEL_ID = os.getenv("MODEL_ID_1")
PROJECT_ID = os.getenv("PROJECT_ID_1")
API_KEY = os.getenv("API_KEY_1")

# API Endpoints
LOGIN_URL = "https://dms.asite.com/apilogin/"
IAM_TOKEN_URL = "https://iam.cloud.ibm.com/identity/token"

import time
from functools import wraps
import streamlit as st

def function_timer(show_args=False):
    """Decorator to measure and display function execution time"""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            # Start timer
            start_time = time.time()
            
            # Call the original function
            result = func(*args, **kwargs)
            
            # Calculate duration
            duration = time.time() - start_time
            
            # Display timing info
            func_name = func.__name__.replace('_', ' ').title()
            arg_info = ""
            if show_args and args:
                arg_info = f" with args: {args[1:]}"  # Skip self if present
            
            st.info(f"⏱️ {func_name}{arg_info} executed in {duration:.2f} seconds")
            
            return result
        return wrapper
    return decorator

# Login Function
@function_timer()
async def login_to_asite(email, password):
    headers = {"Accept": "application/json", "Content-Type": "application/x-www-form-urlencoded"}
    payload = {"emailId": email, "password": password}
    try:
        response = requests.post(LOGIN_URL, headers=headers, data=payload, verify=certifi.where(), timeout=50)
        if response.status_code == 200:
            try:
                session_id = response.json().get("UserProfile", {}).get("Sessionid")
                if session_id:
                    logger.info(f"Login successful, Session ID: {session_id}")
                    st.session_state.sessionid = session_id
                    st.sidebar.success(f"✅ Login successful, Session ID: {session_id}")
                    return session_id
                else:
                    logger.error("No Session ID found in login response")
                    st.sidebar.error("❌ No Session ID in response")
                    return None
            except json.JSONDecodeError:
                logger.error("JSONDecodeError during login")
                st.sidebar.error("❌ Failed to parse login response")
                return None
        logger.error(f"Login failed: {response.status_code} - {response.text}")
        st.sidebar.error(f"❌ Login failed: {response.status_code}")
        return None
    except Exception as e:
        logger.error(f"Error during login: {str(e)}")
        st.sidebar.error(f"❌ Login error: {str(e)}")
        return None

# Function to generate access token
@function_timer()
@retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=2, min=10, max=60))
def get_access_token(api_key):
    headers = {"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"}
    data = {"grant_type": "urn:ibm:params:oauth:grant-type:apikey", "apikey": api_key}
    response = requests.post(IAM_TOKEN_URL, headers=headers, data=data, verify=certifi.where(), timeout=50)
    try:
        if response.status_code == 200:
            token_info = response.json()
            logger.info("Access token generated successfully")
            return token_info['access_token']
        else:
            logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
            st.error(f"❌ Failed to get access token: {response.status_code} - {response.text}")
            raise Exception("Failed to get access token")
    except Exception as e:
        logger.error(f"Exception getting access token: {str(e)}")
        st.error(f"❌ Error getting access token: {str(e)}")
        return None

# Initialize COS client
@function_timer()
@retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=1, min=4, max=10))
def initialize_cos_client():
    try:
        logger.info("Attempting to initialize COS client...")
        cos_client = ibm_boto3.client(
            's3',
            ibm_api_key_id=COS_API_KEY,
            ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
            config=Config(
                signature_version='oauth',
                connect_timeout=180,
                read_timeout=180,
                retries={'max_attempts': 15}
            ),
            endpoint_url=COS_ENDPOINT
        )
        logger.info("COS client initialized successfully")
        return cos_client
    except Exception as e:
        logger.error(f"Error initializing COS client: {str(e)}")
        st.error(f"❌ Error initializing COS client: {str(e)}")
        raise

async def validate_session():
    url = "https://dmsak.asite.com/api/workspace/workspacelist"
    headers = {'Cookie': f'ASessionID={st.session_state.sessionid}'}
    async with aiohttp.ClientSession() as session:
        async with session.get(url, headers=headers) as response:
            if response.status == 200:
                logger.info("Session validated successfully")
                return True
            else:
                logger.error(f"Session validation failed: {response.status} - {await response.text()}")
                return False

async def refresh_session_if_needed():
    if 'sessionid' not in st.session_state or not st.session_state.sessionid:
        logger.warning("No session ID found in session state, attempting login")
        new_session_id = await login_to_asite(os.getenv("ASITE_EMAIL"), os.getenv("ASITE_PASSWORD"))
        if new_session_id:
            st.session_state.sessionid = new_session_id
            return new_session_id
        else:
            raise Exception("Failed to establish initial session")

    if not await validate_session():
        logger.info("Session invalid, attempting to refresh")
        new_session_id = await login_to_asite(os.getenv("ASITE_EMAIL"), os.getenv("ASITE_PASSWORD"))
        if new_session_id:
            st.session_state.sessionid = new_session_id
            logger.info(f"Session refreshed successfully, new Session ID: {new_session_id}")
            return new_session_id
        else:
            raise Exception("Failed to refresh session")
    logger.info("Session is valid, no refresh needed")
    return st.session_state.sessionid

# Fetch Workspace ID
@function_timer()
async def GetWorkspaceID():
    await refresh_session_if_needed()
    url = "https://dmsak.asite.com/api/workspace/workspacelist"
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.get(url, headers=headers)
    st.session_state.workspaceid = response.json()['asiteDataList']['workspaceVO'][4]['Workspace_Id']
    st.write(f"Workspace ID: {st.session_state.workspaceid}")

# Fetch Project IDs
@function_timer()
async def GetProjectId():
    await refresh_session_if_needed()
    url = f"https://adoddleak.asite.com/commonapi/qaplan/getQualityPlanList;searchCriteria={{'criteria': [{{'field': 'planCreationDate','operator': 6,'values': ['11-Mar-2025']}}], 'projectId': {str(st.session_state.workspaceid)}, 'recordLimit': 1000, 'recordStart': 1}}"
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()
        logger.info(f"GetProjectId response: {json.dumps(data, indent=2)}")
        if 'data' not in data or not data['data']:
            st.error("❌ No project data found in GetProjectId response")
            logger.error("No project data found in GetProjectId response")
            return
        st.session_state.veridia_Common_Area_Finishing = data['data'][2]['planId']
        st.session_state.veridia_lift = data['data'][5]['planId']
        st.session_state.veridia_external_development = data['data'][3]['planId']
        st.session_state.veridia_finishing = data['data'][4]['planId']
        st.session_state.veridia_structure = data['data'][6]['planId']
        logger.info(f"Veridia Lift planId: {st.session_state.veridia_lift}")
        st.write(f"Veridia - Lift Project ID: {data['data'][5]['planId']}")
        st.write(f"Veridia - Common Area Finishing Project ID: {data['data'][2]['planId']}")
        st.write(f"Veridia - External Development Project ID: {data['data'][3]['planId']}")
        st.write(f"Veridia Finishing Project ID: {data['data'][4]['planId']}")
        st.write(f"Veridia Structure Project ID: {data['data'][6]['planId']}")
    except Exception as e:
        st.error(f"❌ Error fetching Project IDs: {str(e)}")
        logger.error(f"Error fetching Project IDs: {str(e)}")

# Asynchronous Fetch Function with Retry Logic
@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=10))
async def fetch_data(session, url, headers, email=None, password=None):
    try:
        logger.info(f"Fetching data from URL: {url}")
        async with session.get(url, headers=headers) as response:
            content_type = response.headers.get('Content-Type', '')
            logger.info(f"Response Content-Type: {content_type}")
            if response.status == 200:
                if 'application/json' in content_type:
                    return await response.json()
                else:
                    text = await response.text()
                    logger.error(f"Unexpected Content-Type: {content_type}, Response: {text[:500]}")
                    raise ValueError(f"Unexpected Content-Type: {content_type}, expected application/json")
            elif response.status == 204:
                logger.info("No content returned (204)")
                return None
            else:
                text = await response.text()
                logger.error(f"Error fetching data: {response.status} - {text[:500]}")
                if response.status == 401:
                    raise Exception("Unauthorized: Session may have expired")
                raise Exception(f"Error fetching data: {response.status} - {text[:500]}")
    except Exception as e:
        logger.error(f"Fetch failed: {str(e)}")
        raise

# Fetch All Data with Async
@function_timer()
async def GetAllDatas():
    record_limit = 1000
    all_finishing_data = []
    all_structure_data = []
    all_external_data = []
    all_lift_data = []
    all_common_area_finishing = []

    # Ensure session is valid before starting
    await refresh_session_if_needed()
    headers = {'Cookie': f'ASessionID={st.session_state.sessionid}'}

    async with aiohttp.ClientSession() as session:
        # Fetch Veridia Finishing data
        start_record = 1
        st.write("Fetching Veridia Finishing data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_finishing}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                # Refresh session before each major fetch to ensure validity
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Finishing data available (204)")
                    break
                if 'associationList' in data and data['associationList']:
                    all_finishing_data.extend(data['associationList'])
                else:
                    all_finishing_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_finishing_data[-record_limit:])} Finishing records (Total: {len(all_finishing_data)})")
                if len(all_finishing_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f"❌ Error fetching Finishing data: {str(e)}")
                logger.error(f"Finishing data fetch failed: {str(e)}")
                break

        # Fetch Veridia Structure data
        start_record = 1
        st.write("Fetching Veridia Structure data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure data available (204)")
                    break
                if 'associationList' in data and data['associationList']:
                    all_structure_data.extend(data['associationList'])
                else:
                    all_structure_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_structure_data[-record_limit:])} Structure records (Total: {len(all_structure_data)})")
                if len(all_structure_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f"❌ Error fetching Structure data: {str(e)}")
                logger.error(f"Structure data fetch failed: {str(e)}")
                break

        # Fetch Veridia External Development data
        start_record = 1
        st.write("Fetching Veridia External Development data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_external_development}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more External Development data available (204)")
                    break
                if 'associationList' in data and data['associationList']:
                    all_external_data.extend(data['associationList'])
                else:
                    all_external_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_external_data[-record_limit:])} External Development records (Total: {len(all_external_data)})")
                if len(all_external_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f"❌ Error fetching External Development data: {str(e)}")
                logger.error(f"External Development data fetch failed: {str(e)}")
                break

        # Fetch Veridia Lift data
        start_record = 1
        st.write("Fetching Veridia Lift data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_lift}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Lift data available (204)")
                    break
                if 'associationList' in data and data['associationList']:
                    all_lift_data.extend(data['associationList'])
                else:
                    all_lift_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_lift_data[-record_limit:])} Lift records (Total: {len(all_lift_data)})")
                if len(all_lift_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f"❌ Error fetching Lift data: {str(e)}")
                logger.error(f"Lift data fetch failed: {str(e)}")
                break

        # Fetch Veridia Common Area Finishing data
        start_record = 1
        st.write("Fetching Veridia Common Area Finishing data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_Common_Area_Finishing}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Common Area Finishing data available (204)")
                    break
                if 'associationList' in data and data['associationList']:
                    all_common_area_finishing.extend(data['associationList'])
                else:
                    all_common_area_finishing.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_common_area_finishing[-record_limit:])} Common Area Finishing records (Total: {len(all_common_area_finishing)})")
                if len(all_common_area_finishing[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f"❌ Error fetching Common Area Finishing data: {str(e)}")
                logger.error(f"Common Area Finishing data fetch failed: {str(e)}")
                break

    df_finishing = pd.DataFrame(all_finishing_data)
    df_structure = pd.DataFrame(all_structure_data)
    df_external = pd.DataFrame(all_external_data)
    df_lift = pd.DataFrame(all_lift_data)
    df_common_area = pd.DataFrame(all_common_area_finishing)
    desired_columns = ['activitySeq', 'qiLocationId']
    if 'statusName' in df_finishing.columns:
        desired_columns.append('statusName')
    elif 'statusColor' in df_finishing.columns:
        desired_columns.append('statusColor')
        status_mapping = {'#4CAF50': 'Completed', '#4CB0F0': 'Not Started', '#4C0F0': 'Not Started'}
        df_finishing['statusName'] = df_finishing['statusColor'].map(status_mapping).fillna('Unknown')
        df_structure['statusName'] = df_structure['statusColor'].map(status_mapping).fillna('Unknown')
        df_external['statusName'] = df_external['statusColor'].map(status_mapping).fillna('Unknown')
        df_lift['statusName'] = df_lift['statusColor'].map(status_mapping).fillna('Unknown')
        df_common_area['statusName'] = df_common_area['statusColor'].map(status_mapping).fillna('Unknown')
        desired_columns.append('statusName')
    else:
        st.error("❌ Neither statusName nor statusColor found in data!")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    veridia_finishing = df_finishing[desired_columns]
    veridia_structure = df_structure[desired_columns]
    veridia_external = df_external[desired_columns]
    veridia_lift = df_lift[desired_columns]
    veridia_common_area = df_common_area[desired_columns]

    st.write(f"VERIDIA FINISHING ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(veridia_finishing)}")
    st.write(veridia_finishing)
    st.write(f"VERIDIA STRUCTURE ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(veridia_structure)}")
    st.write(veridia_structure)
    st.write(f"VERIDIA EXTERNAL DEVELOPMENT ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(veridia_external)}")
    st.write(veridia_external)
    st.write(f"VERIDIA LIFT ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(veridia_lift)}")
    st.write(veridia_lift)
    st.write(f"VERIDIA COMMON AREA FINISHING ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(veridia_common_area)}")
    st.write(veridia_common_area)

    return veridia_finishing, veridia_structure, veridia_external, veridia_lift, veridia_common_area


# Fetch Activity Data with Async
@function_timer()
async def Get_Activity():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    all_finishing_activity_data = []
    all_structure_activity_data = []
    all_external_activity_data = []
    all_lift_activity_data = []
    all_common_area_activity_data = []

    # Ensure session is valid before starting
    await refresh_session_if_needed()

    async with aiohttp.ClientSession() as session:
        # Fetch Veridia Finishing Activity data
        start_record = 1
        st.write("Fetching Activity data for Veridia Finishing...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_finishing}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Finishing Activity data available (204)")
                    break
                if 'activityList' in data and data['activityList']:
                    all_finishing_activity_data.extend(data['activityList'])
                else:
                    all_finishing_activity_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_finishing_activity_data[-record_limit:])} Finishing Activity records (Total: {len(all_finishing_activity_data)})")
                if len(all_finishing_activity_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f"❌ Error fetching Finishing Activity data: {str(e)}")
                logger.error(f"Finishing Activity fetch failed: {str(e)}")
                break

        # Fetch Veridia Structure Activity data
        start_record = 1
        st.write("Fetching Activity data for Veridia Structure...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure Activity data available (204)")
                    break
                if 'activityList' in data and data['activityList']:
                    all_structure_activity_data.extend(data['activityList'])
                else:
                    all_structure_activity_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_structure_activity_data[-record_limit:])} Structure Activity records (Total: {len(all_structure_activity_data)})")
                if len(all_structure_activity_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f"❌ Error fetching Structure Activity data: {str(e)}")
                logger.error(f"Structure Activity fetch failed: {str(e)}")
                break

        # Fetch Veridia External Development Activity data
        start_record = 1
        st.write("Fetching Activity data for Veridia External Development...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_external_development}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more External Development Activity data available (204)")
                    break
                if 'activityList' in data and data['activityList']:
                    all_external_activity_data.extend(data['activityList'])
                else:
                    all_external_activity_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_external_activity_data[-record_limit:])} External Development Activity records (Total: {len(all_external_activity_data)})")
                if len(all_external_activity_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f"❌ Error fetching External Development Activity data: {str(e)}")
                logger.error(f"External Development Activity fetch failed: {str(e)}")
                break

        # Fetch Veridia Lift Activity data
        start_record = 1
        st.write("Fetching Activity data for Veridia Lift...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_lift}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                logger.info(f"Fetching Lift Activity data from URL: {url}")
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Lift Activity data available (204)")
                    break
                if 'activityList' in data and data['activityList']:
                    all_lift_activity_data.extend(data['activityList'])
                else:
                    all_lift_activity_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_lift_activity_data[-record_limit:])} Lift Activity records (Total: {len(all_lift_activity_data)})")
                if len(all_lift_activity_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f"❌ Error fetching Lift Activity data: {str(e)}")
                logger.error(f"Lift Activity fetch failed: {str(e)}")
                all_lift_activity_data = []  # Fallback to empty list
                break

        # Fetch Veridia Common Area Finishing Activity data
        start_record = 1
        st.write("Fetching Activity data for Veridia Common Area Finishing...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_Common_Area_Finishing}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Common Area Finishing Activity data available (204)")
                    break
                if 'activityList' in data and data['activityList']:
                    all_common_area_activity_data.extend(data['activityList'])
                else:
                    all_common_area_activity_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_common_area_activity_data[-record_limit:])} Common Area Finishing Activity records (Total: {len(all_common_area_activity_data)})")
                if len(all_common_area_activity_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f"❌ Error fetching Common Area Finishing Activity data: {str(e)}")
                logger.error(f"Common Area Finishing Activity fetch failed: {str(e)}")
                break

    def safe_select(df, cols):
        if df.empty:
            return pd.DataFrame(columns=cols)
        missing = [col for col in cols if col not in df.columns]
        if missing:
            logger.warning(f"Missing columns in activity data: {missing}")
            for col in missing:
                df[col] = None
        return df[cols]

    finishing_activity_data = safe_select(pd.DataFrame(all_finishing_activity_data), ['activityName', 'activitySeq', 'formTypeId'])
    structure_activity_data = safe_select(pd.DataFrame(all_structure_activity_data), ['activityName', 'activitySeq', 'formTypeId'])
    external_activity_data = safe_select(pd.DataFrame(all_external_activity_data), ['activityName', 'activitySeq', 'formTypeId'])
    lift_activity_data = safe_select(pd.DataFrame(all_lift_activity_data), ['activityName', 'activitySeq', 'formTypeId'])
    common_area_activity_data = safe_select(pd.DataFrame(all_common_area_activity_data), ['activityName', 'activitySeq', 'formTypeId'])

    st.write("VERIDIA FINISHING ACTIVITY DATA (activityName, activitySeq, formTypeId)")
    st.write(f"Total records: {len(finishing_activity_data)}")
    st.write(finishing_activity_data)
    st.write("VERIDIA STRUCTURE ACTIVITY DATA (activityName, activitySeq, formTypeId)")
    st.write(f"Total records: {len(structure_activity_data)}")
    st.write(structure_activity_data)
    st.write("VERIDIA EXTERNAL DEVELOPMENT ACTIVITY DATA (activityName, activitySeq, formTypeId)")
    st.write(f"Total records: {len(external_activity_data)}")
    st.write(external_activity_data)
    st.write("VERIDIA LIFT ACTIVITY DATA (activityName, activitySeq, formTypeId)")
    st.write(f"Total records: {len(lift_activity_data)}")
    st.write(lift_activity_data)
    st.write("VERIDIA COMMON AREA FINISHING ACTIVITY DATA (activityName, activitySeq, formTypeId)")
    st.write(f"Total records: {len(common_area_activity_data)}")
    st.write(common_area_activity_data)

    return finishing_activity_data, structure_activity_data, external_activity_data, lift_activity_data, common_area_activity_data

# Fetch Location/Module Data with Async
@function_timer()
async def Get_Location():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    all_finishing_location_data = []
    all_structure_location_data = []
    all_external_location_data = []
    all_lift_location_data = []
    all_common_area_location_data = []

    # Ensure session is valid before starting
    await refresh_session_if_needed()

    async with aiohttp.ClientSession() as session:
        # Fetch Veridia Finishing Location/Module data
        start_record = 1
        total_records_fetched = 0
        st.write("Fetching Veridia Finishing Location/Module data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_finishing}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Finishing Location data available (204)")
                    break
                if isinstance(data, list):
                    location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
                                   for item in data if isinstance(item, dict)]
                    all_finishing_location_data.extend(location_data)
                    total_records_fetched = len(all_finishing_location_data)
                    st.write(f"Fetched {len(location_data)} Finishing Location records (Total: {total_records_fetched})")
                elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                    location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
                                   for loc in data['locationList']]
                    all_finishing_location_data.extend(location_data)
                    total_records_fetched = len(all_finishing_location_data)
                    st.write(f"Fetched {len(location_data)} Finishing Location records (Total: {total_records_fetched})")
                else:
                    st.warning(f"No 'locationList' in Finishing Location data or empty list.")
                    break
                if len(location_data) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f"❌ Error fetching Finishing Location data: {str(e)}")
                logger.error(f"Finishing Location fetch failed: {str(e)}")
                break

        # Fetch Veridia Structure Location/Module data
        start_record = 1
        total_records_fetched = 0
        st.write("Fetching Veridia Structure Location/Module data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure Location data available (204)")
                    break
                if isinstance(data, list):
                    location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
                                   for item in data if isinstance(item, dict)]
                    all_structure_location_data.extend(location_data)
                    total_records_fetched = len(all_structure_location_data)
                    st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                    location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
                                   for loc in data['locationList']]
                    all_structure_location_data.extend(location_data)
                    total_records_fetched = len(all_structure_location_data)
                    st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                else:
                    st.warning(f"No 'locationList' in Structure Location data or empty list.")
                    break
                if len(location_data) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f"❌ Error fetching Structure Location data: {str(e)}")
                logger.error(f"Structure Location fetch failed: {str(e)}")
                break

        # Fetch Veridia External Development Location/Module data
        start_record = 1
        total_records_fetched = 0
        st.write("Fetching Veridia External Development Location/Module data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_external_development}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more External Development Location data available (204)")
                    break
                if isinstance(data, list):
                    location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
                                   for item in data if isinstance(item, dict)]
                    all_external_location_data.extend(location_data)
                    total_records_fetched = len(all_external_location_data)
                    st.write(f"Fetched {len(location_data)} External Development Location records (Total: {total_records_fetched})")
                elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                    location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
                                   for loc in data['locationList']]
                    all_external_location_data.extend(location_data)
                    total_records_fetched = len(all_external_location_data)
                    st.write(f"Fetched {len(location_data)} External Development Location records (Total: {total_records_fetched})")
                else:
                    st.warning(f"No 'locationList' in External Development Location data or empty list.")
                    break
                if len(location_data) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f"❌ Error fetching External Development Location data: {str(e)}")
                logger.error(f"External Development Location fetch failed: {str(e)}")
                break

        # Fetch Veridia Lift Location/Module data
        start_record = 1
        total_records_fetched = 0
        st.write("Fetching Veridia Lift Location/Module data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_lift}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                logger.info(f"Fetching Lift Location data from URL: {url}")
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Lift Location data available (204)")
                    break
                if isinstance(data, list):
                    location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
                                   for item in data if isinstance(item, dict)]
                    all_lift_location_data.extend(location_data)
                    total_records_fetched = len(all_lift_location_data)
                    st.write(f"Fetched {len(location_data)} Lift Location records (Total: {total_records_fetched})")
                elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                    location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
                                   for loc in data['locationList']]
                    all_lift_location_data.extend(location_data)
                    total_records_fetched = len(all_lift_location_data)
                    st.write(f"Fetched {len(location_data)} Lift Location records (Total: {total_records_fetched})")
                else:
                    st.warning(f"No 'locationList' in Lift Location data or empty list.")
                    break
                if len(location_data) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f"❌ Error fetching Lift Location data: {str(e)}")
                logger.error(f"Lift Location fetch failed: {str(e)}")
                all_lift_location_data = []  # Fallback to empty list
                break

        # Fetch Veridia Common Area Finishing Location/Module data
        start_record = 1
        total_records_fetched = 0
        st.write("Fetching Veridia Common Area Finishing Location/Module data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.veridia_Common_Area_Finishing}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Common Area Finishing Location data available (204)")
                    break
                if isinstance(data, list):
                    location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
                                   for item in data if isinstance(item, dict)]
                    all_common_area_location_data.extend(location_data)
                    total_records_fetched = len(all_common_area_location_data)
                    st.write(f"Fetched {len(location_data)} Common Area Finishing Location records (Total: {total_records_fetched})")
                elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                    location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
                                   for loc in data['locationList']]
                    all_common_area_location_data.extend(location_data)
                    total_records_fetched = len(all_common_area_location_data)
                    st.write(f"Fetched {len(location_data)} Common Area Finishing Location records (Total: {total_records_fetched})")
                else:
                    st.warning(f"No 'locationList' in Common Area Finishing Location data or empty list.")
                    break
                if len(location_data) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f"❌ Error fetching Common Area Finishing Location data: {str(e)}")
                logger.error(f"Common Area Finishing Location fetch failed: {str(e)}")
                break

    finishing_df = pd.DataFrame(all_finishing_location_data)
    structure_df = pd.DataFrame(all_structure_location_data)
    external_df = pd.DataFrame(all_external_location_data)
    lift_df = pd.DataFrame(all_lift_location_data)
    common_area_df = pd.DataFrame(all_common_area_location_data)

    # Validate name field
    if 'name' in finishing_df.columns and finishing_df['name'].isna().all():
        st.error("❌ All 'name' values in Finishing Location data are missing or empty!")
    if 'name' in structure_df.columns and structure_df['name'].isna().all():
        st.error("❌ All 'name' values in Structure Location data are missing or empty!")
    if 'name' in external_df.columns and external_df['name'].isna().all():
        st.error("❌ All 'name' values in External Development Location data are missing or empty!")
    if 'name' in lift_df.columns and lift_df['name'].isna().all():
        st.error("❌ All 'name' values in Lift Location data are missing or empty!")
    if 'name' in common_area_df.columns and common_area_df['name'].isna().all():
        st.error("❌ All 'name' values in Common Area Finishing Location data are missing or empty!")

    st.write("VERIDIA FINISHING LOCATION/MODULE DATA")
    st.write(f"Total records: {len(finishing_df)}")
    st.write(finishing_df)
    st.write("VERIDIA STRUCTURE LOCATION/MODULE DATA")
    st.write(f"Total records: {len(structure_df)}")
    st.write(structure_df)
    st.write("VERIDIA EXTERNAL DEVELOPMENT LOCATION/MODULE DATA")
    st.write(f"Total records: {len(external_df)}")
    st.write(external_df)
    st.write("VERIDIA LIFT LOCATION/MODULE DATA")
    st.write(f"Total records: {len(lift_df)}")
    st.write(lift_df)
    st.write("VERIDIA COMMON AREA FINISHING LOCATION/MODULE DATA")
    st.write(f"Total records: {len(common_area_df)}")
    st.write(common_area_df)

    st.session_state.finishing_location_data = finishing_df
    st.session_state.structure_location_data = structure_df
    st.session_state.external_location_data = external_df
    st.session_state.lift_location_data = lift_df
    st.session_state.common_area_location_data = common_area_df

    return finishing_df, structure_df, external_df, lift_df, common_area_df


# Process individual chunk
@function_timer()
def process_chunk(chunk, chunk_idx, dataset_name, location_df):
    logger.info(f"Starting thread for {dataset_name} Chunk {chunk_idx + 1}")
    generated_text = format_chunk_locally(chunk, chunk_idx, len(chunk), dataset_name, location_df)
    logger.info(f"Completed thread for {dataset_name} Chunk {chunk_idx + 1}")
    return generated_text, chunk_idx

# Process data with manual counting
@function_timer()
def process_manually(analysis_df, total, dataset_name, chunk_size=1000, max_workers=4):
    if analysis_df.empty:
        st.warning(f"No completed activities found for {dataset_name}.")
        return {"towers": {}, "total": 0}

    unique_activities = analysis_df['activityName'].unique()
    logger.info(f"Unique activities in {dataset_name} dataset: {list(unique_activities)}")
    logger.info(f"Total records in {dataset_name} dataset: {len(analysis_df)}")

    st.write(f"Saved Veridia {dataset_name} data to veridia_{dataset_name.lower()}_data.json")
    chunks = [analysis_df[i:i + chunk_size] for i in range(0, len(analysis_df), chunk_size)]

    location_df = (
        st.session_state.finishing_location_data if dataset_name.lower() == "finishing"
        else st.session_state.structure_location_data if dataset_name.lower() == "structure"
        else st.session_state.external_location_data
    )

    chunk_results = {}
    progress_bar = st.progress(0)
    status_text = st.empty()

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_chunk = {
            executor.submit(process_chunk, chunk, idx, dataset_name, location_df): idx
            for idx, chunk in enumerate(chunks)
        }

        completed_chunks = 0
        for future in as_completed(future_to_chunk):
            chunk_idx = future_to_chunk[future]
            try:
                generated_text, idx = future.result()
                chunk_results[idx] = generated_text
                completed_chunks += 1
                progress_percent = completed_chunks / len(chunks)
                progress_bar.progress(progress_percent)
                status_text.text(f"Processed chunk {completed_chunks} of {len(chunks)} ({progress_percent:.1%} complete)")
            except Exception as e:
                logger.error(f"Error processing chunk {chunk_idx + 1} for {dataset_name}: {str(e)}")
                st.error(f"❌ Error processing chunk {chunk_idx + 1}: {str(e)}")

    parsed_data = {}
    for chunk_idx in sorted(chunk_results.keys()):
        generated_text = chunk_results[chunk_idx]
        if not generated_text:
            continue

        current_tower = None
        tower_activities = []
        lines = generated_text.split("\n")

        for line in lines:
            line = line.strip()
            if not line:
                continue

            if line.startswith("Tower:"):
                try:
                    tower_parts = line.split("Tower:", 1)
                    if len(tower_parts) > 1:
                        if current_tower and tower_activities:
                            if current_tower not in parsed_data:
                                parsed_data[current_tower] = []
                            parsed_data[current_tower].extend(tower_activities)
                            tower_activities = []
                        current_tower = tower_parts[1].strip()
                except Exception as e:
                    logger.warning(f"Error parsing Tower line: {line}, error: {str(e)}")
                    if not current_tower:
                        current_tower = f"Unknown Tower {chunk_idx}"

            elif line.startswith("Total Completed Activities:"):
                continue
            elif not line.strip().startswith("activityName"):
                try:
                    parts = re.split(r'\s{2,}', line.strip())
                    if len(parts) >= 2:
                        activity_name = ' '.join(parts[:-1]).strip()
                        count_str = parts[-1].strip()
                        count_match = re.search(r'\d+', count_str)
                        if count_match:
                            count = int(count_match.group())
                            if current_tower:
                                tower_activities.append({
                                    "activityName": activity_name,
                                    "completedCount": count
                                })
                    else:
                        match = re.match(r'^\s*(.+?)\s+(\d+)$', line)
                        if match and current_tower:
                            activity_name = match.group(1).strip()
                            count = int(match.group(2).strip())
                            tower_activities.append({
                                "activityName": activity_name,
                                "completedCount": count
                            })
                except (ValueError, IndexError) as e:
                    logger.warning(f"Skipping malformed activity line: {line}, error: {str(e)}")

        if current_tower and tower_activities:
            if current_tower not in parsed_data:
                parsed_data[current_tower] = []
            parsed_data[current_tower].extend(tower_activities)

    aggregated_data = {}
    for tower_name, activities in parsed_data.items():
        tower_short_name = tower_name.split('/')[1] if '/' in tower_name else tower_name
        if tower_short_name not in aggregated_data:
            aggregated_data[tower_short_name] = {}
        
        for activity in activities:
            name = activity.get("activityName", "Unknown")
            count = activity.get("completedCount", 0)
            if name in aggregated_data[tower_short_name]:
                aggregated_data[tower_short_name][name] += count
            else:
                aggregated_data[tower_short_name][name] = count

    return {"towers": aggregated_data, "total": total}

# Local formatting function for manual counting
@function_timer()
def format_chunk_locally(chunk, chunk_idx, chunk_size, dataset_name, location_df):
    towers_data = {}
    
    for _, row in chunk.iterrows():
        tower_name = row['tower_name']
        activity_name = row['activityName']
        count = int(row['CompletedCount'])
        
        if tower_name not in towers_data:
            towers_data[tower_name] = []
            
        towers_data[tower_name].append({
            "activityName": activity_name,
            "completedCount": count
        })
    
    output = ""
    total_activities = 0
    
    for tower_name, activities in sorted(towers_data.items()):
        output += f"Tower: {tower_name}\n"
        output += "activityName            CompletedCount\n"
        activity_dict = {}
        for activity in activities:
            name = activity['activityName']
            count = activity['completedCount']
            activity_dict[name] = activity_dict.get(name, 0) + count
        for name, count in sorted(activity_dict.items()):
            output += f"{name:<30} {count}\n"
            total_activities += count
    
    output += f"Total Completed Activities: {total_activities}"
    return output


@function_timer()
def process_data(df, activity_df, location_df, dataset_name, use_module_hierarchy_for_finishing=False):
    # Filter completed activities
    completed = df[df['statusName'] == 'Completed'].copy()
    
    # Define expected Asite activities for count_table
    asite_activities = [
        "Wall Conducting", "Plumbing Works", "POP & Gypsum Plaster", "Wiring & Switch Socket",
        "Slab Conducting", "Electrical Cable", "Door/Window Frame", "Waterproofing - Sunken",
        "Wall Tile", "Floor Tile", "Door/Window Shutter", "Shuttering", "Reinforcement",
        "Sewer Line", "Rain Water/Storm Line", "Granular Sub-base", "WMM",
        "Saucer drain/Paver block", "Kerb Stone", "Concreting"
    ]
    
    # Initialize empty count_table
    count_table = pd.DataFrame({'Count': [0] * len(asite_activities)}, index=asite_activities)
    
    if completed.empty:
        logger.warning(f"No completed activities found in {dataset_name} data.")
        return pd.DataFrame(), 0, count_table

    # Merge with location and activity data
    completed = completed.merge(location_df[['qiLocationId', 'name']], on='qiLocationId', how='left')
    completed = completed.merge(activity_df[['activitySeq', 'activityName']], on='activitySeq', how='left')

    if 'qiActivityId' not in completed.columns:
        completed['qiActivityId'] = completed['qiLocationId'].astype(str) + '$$' + completed['activitySeq'].astype(str)

    if completed['name'].isna().all():
        logger.error(f"All 'name' values are missing in {dataset_name} data after merge!")
        st.error(f"❌ All 'name' values are missing in {dataset_name} data after merge! Check location data.")
        completed['name'] = 'Unknown'
    else:
        completed['name'] = completed['name'].fillna('Unknown')

    def normalize_activity_name(name):
        """
        Normalize activity names to fix common typos and ensure consistency with expected categories.
        """
        if not isinstance(name, str):
            return name
        typo_corrections = {
            "Wall Conduting": "Wall Conducting",
            "Slab conduting": "Slab Conducting",
            "WallTile": "Wall Tile",
            "FloorTile": "Floor Tile",
            "wall tile": "Wall Tile",
            "floor tile": "Floor Tile",
            "DoorWindowFrame": "Door/Window Frame",
            "DoorWindowShutter": "Door/Window Shutter",
            "Second Roof Slab": "Roof Slab",
            "First Roof Slab": "Roof Slab",
            "Roof slab": "Roof Slab",
            "Beam": "Beam",
            "Column": "Column",
            "Reinforcement": "Reinforcement",
            "Shuttering": "Shuttering",
            "Concreting": "Concreting",
            "DeShuttering": "De-Shuttering"
        }
        for typo, correct in typo_corrections.items():
            if name.lower() == typo.lower():
                return correct
        return name

    completed['activityName'] = completed['activityName'].apply(normalize_activity_name).fillna('Unknown')

    # Build location path dictionaries
    parent_child_dict = dict(zip(location_df['qiLocationId'], location_df['qiParentId']))
    name_dict = dict(zip(location_df['qiLocationId'], location_df['name']))

    def get_full_path(location_id):
        """
        Construct full location path (e.g., Quality/Tower 1/Module 1/North/101).
        """
        path = []
        current_id = location_id
        max_depth = 10
        depth = 0
        
        while current_id and depth < max_depth:
            if current_id not in parent_child_dict or current_id not in name_dict:
                logger.warning(f"Location ID {current_id} not found in parent_child_dict or name_dict. Path so far: {path}")
                break
            
            parent_id = parent_child_dict.get(current_id)
            name = name_dict.get(current_id, "Unknown")
            
            if not parent_id:
                if name != "Quality":
                    path.append(name)
                    path.append("Quality")
                else:
                    path.append(name)
                break
            
            path.append(name)
            current_id = parent_id
            depth += 1
        
        if depth >= max_depth:
            logger.warning(f"Max depth reached while computing path for location_id {location_id}. Possible circular reference. Path: {path}")
        
        if not path:
            logger.warning(f"No path constructed for location_id {location_id}. Using 'Unknown'.")
            return "Unknown"
        
        full_path = '/'.join(reversed(path))
        logger.debug(f"Full path for location_id {location_id}: {full_path}")
        return full_path

    completed['full_path'] = completed['qiLocationId'].apply(get_full_path)

    # Debug: Log unique full_path values before filtering
    logger.debug(f"All unique full_path values in {dataset_name} dataset BEFORE filtering:")
    unique_paths = completed['full_path'].unique()
    for path in sorted(unique_paths):
        logger.debug(f"  Path: {path}")
    
    # Log tower distribution before filtering
    completed['temp_tower_name'] = completed['full_path'].apply(lambda x: x.split('/')[1] if len(x.split('/')) > 1 else x)
    tower_counts_before = completed['temp_tower_name'].value_counts()
    logger.debug(f"Tower distribution BEFORE filtering in {dataset_name}:")
    for tower, count in tower_counts_before.items():
        logger.debug(f"  {tower}: {count} records")

    # Define filtering functions
    def has_flat_number(full_path):
        parts = full_path.split('/')
        last_part = parts[-1]
        match = re.match(r'^\d+(?:(?:\s*\(LL\))|(?:\s*\(UL\))|(?:\s*LL)|(?:\s*UL))?$', last_part)
        return bool(match)

    def is_roof_slab(full_path):
        parts = full_path.split('/')
        last_part = parts[-1].lower()
        return last_part.endswith('roof slab')

    def is_roof_slab_only(full_path):
        """
        Strict filtering for structure dataset - ONLY roof slab locations.
        """
        parts = full_path.split('/')
        last_part = parts[-1].lower()
        
        # Only include locations that contain 'roof slab' (case insensitive)
        return 'roof slab' in last_part

    # Apply filtering based on dataset
    if dataset_name.lower() == 'structure':
        # For Structure: ONLY include locations with "roof slab" in the name
        logger.debug(f"Applying STRICT roof slab filtering for {dataset_name} dataset")
        completed_before_filter = len(completed)
        
        # Log all unique paths before filtering to see what we have
        logger.debug(f"All unique paths before roof slab filtering:")
        for path in sorted(completed['full_path'].unique()):
            logger.debug(f"  {path}")
        
        completed = completed[completed['full_path'].apply(is_roof_slab_only)]
        completed_after_filter = len(completed)
        logger.debug(f"Roof slab filtering: {completed_before_filter} -> {completed_after_filter} records")
        
        # Log which paths passed the filter
        if not completed.empty:
            logger.debug(f"Paths that passed roof slab filtering:")
            for path in sorted(completed['full_path'].unique()):
                logger.debug(f"  ✓ {path}")
        else:
            logger.warning(f"No paths contain 'roof slab' in {dataset_name} dataset")
            logger.debug("Checking for similar patterns...")
            all_paths = df[df['statusName'] == 'Completed']['qiLocationId'].apply(get_full_path).unique()
            roof_related = [path for path in all_paths if 'roof' in path.lower() or 'slab' in path.lower()]
            if roof_related:
                logger.debug("Found paths containing 'roof' or 'slab':")
                for path in sorted(roof_related):
                    logger.debug(f"  {path}")
        
        if completed.empty:
            logger.warning(f"No completed activities with 'roof slab' locations found in {dataset_name} data after filtering.")
            return pd.DataFrame(), 0, count_table
    else:
        # For other datasets (Finishing, Lift, External Development, Common Area): Filter for flat numbers
        completed = completed[completed['full_path'].apply(has_flat_number)]
        if completed.empty:
            logger.warning(f"No completed activities with flat numbers found in {dataset_name} data after filtering.")
            return pd.DataFrame(), 0, count_table

    # Log tower distribution after filtering
    completed['temp_tower_name'] = completed['full_path'].apply(lambda x: x.split('/')[1] if len(x.split('/')) > 1 else x)
    tower_counts_after = completed['temp_tower_name'].value_counts()
    logger.debug(f"Tower distribution AFTER filtering in {dataset_name}:")
    for tower, count in tower_counts_after.items():
        logger.debug(f"  {tower}: {count} records")
    completed = completed.drop(columns=['temp_tower_name'])

    def get_tower_name(full_path):
        """
        Extract tower name, splitting Tower 4 into 4(A) and 4(B) based on module.
        """
        parts = full_path.split('/')
        if len(parts) < 2:
            return full_path
        
        tower = parts[1]
        if tower == "Tower 4" and len(parts) > 2:
            module = parts[2]
            module_number = module.replace("Module ", "").strip()
            try:
                module_num = int(module_number)
                if 1 <= module_num <= 4:
                    return "Tower 4(B)"
                elif 5 <= module_num <= 8:
                    return "Tower 4(A)"
            except ValueError:
                logger.warning(f"Could not parse module number from {module} in path {full_path}")
        
        return tower

    completed['tower_name'] = completed['full_path'].apply(get_tower_name)

    # Debug activity names and tower names after filtering
    logger.debug(f"Unique activityName values in completed DataFrame for {dataset_name}:\n{completed['activityName'].unique()}")
    logger.debug(f"Unique tower_name values in completed DataFrame for {dataset_name}:\n{completed['tower_name'].unique()}")
    
    # Log sample paths for each tower to understand the data structure
    for tower in sorted(completed['tower_name'].unique()):
        tower_paths = completed[completed['tower_name'] == tower]['full_path'].unique()[:5]  # Show first 5 paths
        logger.debug(f"Sample paths for {tower}:")
        for path in tower_paths:
            logger.debug(f"  {path}")

    # Create analysis table
    analysis = completed.groupby(['tower_name', 'activityName'])['qiLocationId'].nunique().reset_index(name='CompletedCount')
    analysis = analysis.sort_values(by=['tower_name', 'activityName'], ascending=True)
    total_completed = analysis['CompletedCount'].sum()

    # Populate count_table
    activity_counts = completed.groupby('activityName')['qiLocationId'].nunique().reset_index(name='Count')
    for activity in asite_activities:
        if activity in activity_counts['activityName'].values:
            count_table.loc[activity, 'Count'] = activity_counts[activity_counts['activityName'] == activity]['Count'].iloc[0]

    logger.info(f"Total completed activities for {dataset_name}: {total_completed}")
    logger.info(f"Count table for {dataset_name}:\n{count_table.to_string()}")
    
    # Final debug: Show analysis results by tower
    logger.debug(f"Final analysis results for {dataset_name} by tower:")
    for tower in sorted(analysis['tower_name'].unique()):
        tower_data = analysis[analysis['tower_name'] == tower]
        tower_total = tower_data['CompletedCount'].sum()
        logger.debug(f"  {tower}: {tower_total} total completed activities")
    
    return analysis, total_completed, count_table


# Main analysis function

@function_timer()
def AnalyzeStatusManually(email=None, password=None):
    start_time = time.time()

    if 'sessionid' not in st.session_state:
        st.error("❌ Please log in first!")
        return

    required_data = [
        'veridiafinishing', 'veridiastructure', 'veridiaexternal',
        'veridialift', 'veridiacommonarea',
        'finishing_activity_data', 'structure_activity_data', 'external_activity_data',
        'lift_activity_data', 'common_area_activity_data',
        'finishing_location_data', 'structure_location_data', 'external_location_data',
        'lift_location_data', 'common_area_location_data'
    ]
    
    for data_key in required_data:
        if data_key not in st.session_state:
            st.error(f"❌ Please fetch required data first! Missing: {data_key}")
            return

    try:
        finishing_data = st.session_state.veridiafinishing
        structure_data = st.session_state.veridiastructure
        external_data = st.session_state.veridiaexternal
        lift_data = st.session_state.veridialift
        common_area_data = st.session_state.veridiacommonarea

        finishing_activity = st.session_state.finishing_activity_data
        structure_activity = st.session_state.structure_activity_data
        external_activity = st.session_state.external_activity_data
        lift_activity = st.session_state.lift_activity_data
        common_area_activity = st.session_state.common_area_activity_data

        finishing_locations = st.session_state.finishing_location_data
        structure_locations = st.session_state.structure_location_data
        external_locations = st.session_state.external_location_data
        lift_locations = st.session_state.lift_location_data
        common_area_locations = st.session_state.common_area_location_data
    except KeyError as e:
        st.error(f"❌ Missing session state data: {str(e)}")
        return
    except Exception as e:
        st.error(f"❌ Error retrieving session state data: {str(e)}")
        return

    main_datasets = [
        (finishing_data, "Finishing"),
        (structure_data, "Structure"),
        (external_data, "External Development"),
        (lift_data, "Lift"),
        (common_area_data, "Common Area Finishing")
    ]

    for df, name in main_datasets:
        if not isinstance(df, pd.DataFrame):
            st.error(f"❌ {name} data is not a DataFrame! Type: {type(df)}")
            st.write(f"Content preview: {str(df)[:200]}...")
            return
            
        required_columns = ['statusName', 'qiLocationId', 'activitySeq']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            st.error(f"❌ Missing columns in {name} data: {missing_columns}")
            st.write(f"Available columns: {list(df.columns)}")
            return
    
    location_datasets = [
        (finishing_locations, "Finishing Location"),
        (structure_locations, "Structure Location"),
        (external_locations, "External Development Location"),
        (lift_locations, "Lift Location"),
        (common_area_locations, "Common Area Finishing Location")
    ]

    for df, name in location_datasets:
        if not isinstance(df, pd.DataFrame):
            st.error(f"❌ {name} data is not a DataFrame! Type: {type(df)}")
            return
            
        required_columns = ['qiLocationId', 'name']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            st.error(f"❌ Missing columns in {name} data: {missing_columns}")
            st.write(f"Available columns: {list(df.columns)}")
            return

    activity_datasets = [
        (finishing_activity, "Finishing Activity"),
        (structure_activity, "Structure Activity"),
        (external_activity, "External Development Activity"),
        (lift_activity, "Lift Activity"),
        (common_area_activity, "Common Area Finishing Activity")
    ]

    for df, name in activity_datasets:
        if not isinstance(df, pd.DataFrame):
            st.error(f"❌ {name} data is not a DataFrame! Type: {type(df)}")
            return
            
        required_columns = ['activitySeq', 'activityName']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            st.error(f"❌ Missing columns in {name} data: {missing_columns}")
            st.write(f"Available columns: {list(df.columns)}")
            return

    def normalize_activity_name(name):
        if not isinstance(name, str):
            return name
        typo_corrections = {
            "Wall Conduting": "Wall Conducting",
            "Slab conduting": "Slab Conducting",
        }
        for typo, correct in typo_corrections.items():
            if name.lower() == typo.lower():
                return correct
        return name

    for df in [finishing_activity, structure_activity, external_activity, lift_activity, common_area_activity]:
        df['activityName'] = df['activityName'].apply(normalize_activity_name)

    # Fetch and store COS slab cycle data
    st.write("Fetching COS slab cycle data...")

    # Check if required session state keys exist
    if not all(key in st.session_state for key in ['cos_client', 'bucket_name']):
        st.error("COS client not initialized. Please run 'Initialize and Fetch Data' first.")
        st.session_state['slab_df'] = pd.DataFrame()
    else:
        try:
            # Get the COS client and bucket from session state
            cos_client = st.session_state['cos_client']
            bucket_name = st.session_state['bucket_name']
            
            # Initialize file_list if it doesn't exist or is None
            if 'file_list' not in st.session_state or st.session_state['file_list'] is None:
                st.write("File list not found in session state. Fetching files from COS...")
                try:
                    response = cos_client.list_objects_v2(Bucket=bucket_name, Prefix="Veridia/")
                    if 'Contents' in response:
                        st.session_state['file_list'] = [{'Key': obj['Key']} for obj in response['Contents']]
                    else:
                        st.session_state['file_list'] = []
                        st.warning("No files found in Veridia folder.")
                except Exception as e:
                    st.session_state['file_list'] = []
            
            file_list = st.session_state['file_list']
            
            # Ensure file_list is a list before proceeding
            if not isinstance(file_list, list):
                st.error(f"file_list is not a list. Type: {type(file_list)}")
                st.session_state['slab_df'] = pd.DataFrame()
            else:
                # Call GetSlabReport function if it exists
                try:
                    GetSlabReport()  # Make sure this function is defined and works properly
                    #
                except NameError:
                    st.warning("GetSlabReport function not found. Skipping slab report generation.")
                except Exception as e:
                    st.warning(f"Error calling GetSlabReport: {str(e)}")
                
                # Find slab cycle files
                struct_files = [file for file in file_list if isinstance(file, dict) and 'Key' in file and 'Anti. Slab Cycle' in file['Key']]
            
        except Exception as e:
            st.error(f"❌ Error fetching COS slab cycle data: {str(e)}")
            logging.error(f"Error fetching COS slab cycle data: {str(e)}")
            st.session_state['slab_df'] = pd.DataFrame()

    asite_data = []
    outputs = {}
    for dataset_name, data, activity, location in [
        ("Finishing", finishing_data, finishing_activity, finishing_locations),
        ("Structure", structure_data, structure_activity, structure_locations),
        ("External Development", external_data, external_activity, external_locations),
        ("Lift", lift_data, lift_activity, lift_locations),
        ("Common Area Finishing", common_area_data, common_area_activity, common_area_locations)
    ]:
        try:
            analysis, total, count_table = process_data(data, activity, location, dataset_name)
            if analysis.empty and total == 0:
                logger.warning(f"No valid data processed for {dataset_name}. Skipping analysis.")
                st.warning(f"No completed activities found for {dataset_name}.")
                outputs[dataset_name] = {"towers": {}, "total": 0}
                continue
            output = process_manually(analysis, total, dataset_name)
            outputs[dataset_name] = output
            for tower, activities in output["towers"].items():
                for activity_name, count in activities.items():
                    normalized_name = normalize_activity_name(activity_name)
                    asite_data.append({
                        "Dataset": dataset_name,
                        "Tower": tower,
                        "Activity Name": normalized_name,
                        "Count": count
                    })
        except Exception as e:
            logger.error(f"Error processing {dataset_name} data: {str(e)}")
            st.error(f"❌ Error processing {dataset_name} data: {str(e)}")
            outputs[dataset_name] = {"towers": {}, "total": 0}
            continue

    asite_df = pd.DataFrame(asite_data)

    for dataset_name in ["Finishing", "Structure", "External Development", "Lift", "Common Area Finishing"]:
        output = outputs.get(dataset_name, {"towers": {}, "total": 0})
        st.write(f"### Veridia {dataset_name} Quality Analysis (Completed Activities):")
        if not output["towers"]:
            st.write("No completed activities found.")
            continue
        for tower, activities in output["towers"].items():
            st.write(f"{tower} activityName            CompletedCount")
            for name, count in sorted(activities.items()):
                st.write(f"{'':<11} {name:<23} {count:>14}")
            st.write(f"{'':<11} Total for {tower:<11}: {sum(activities.values()):>14}")
        st.write(f"Total Completed Activities: {output['total']}")

    cos_data = []
    first_fix_counts = {}

    cos_datasets = [
        ('cos_tname_tower4a', 'cos_df_tower4a'),
        ('cos_tname_tower4b', 'cos_df_tower4b'),
        ('cos_tname_tower5', 'cos_df_tower5')
    ]

    for tname_key, tdata_key in cos_datasets:
        if tname_key in st.session_state and tdata_key in st.session_state:
            tname = st.session_state[tname_key]
            tower_data = st.session_state[tdata_key]
            
            if tower_data is not None and isinstance(tower_data, pd.DataFrame):
                tower_data = tower_data.copy()
                tower_data['Actual Finish'] = pd.to_datetime(tower_data['Actual Finish'], errors='coerce')
                tower_data_filtered = tower_data[~pd.isna(tower_data['Actual Finish'])].copy()
                
                first_fix_counts[tname] = {}
                
                for activity in [
                    "EL-First Fix", "UP-First Fix", "CP-First Fix", "C-Gypsum and POP Punning",
                    "EL-Second Fix", "No. of Slab cast", "Electrical", "Installation of doors",
                    "Waterproofing Works", "Wall Tiling", "Floor Tiling", "Sewer Line",
                    "Storm Line", "GSB", "WMM", "Stamp Concrete", "Saucer drain", "Kerb Stone"
                ]:
                    count = len(tower_data_filtered[tower_data_filtered['Activity Name'] == activity])
                    cos_data.append({
                        "Tower": tname,
                        "Activity Name": activity,
                        "Count": count
                    })
                    
                    if activity == "UP-First Fix" or activity == "CP-First Fix":
                        first_fix_counts[tname][activity] = count

    for tname in first_fix_counts:
        up_count = first_fix_counts[tname].get("UP-First Fix", 0)
        cp_count = first_fix_counts[tname].get("CP-First Fix", 0)
        combined_count = up_count + cp_count
        cos_data.append({
            "Tower": tname,
            "Activity Name": "Min. count of UP-First Fix and CP-First Fix",
            "Count": combined_count
        })

    cos_df = pd.DataFrame(cos_data)
    
    logger.info(f"Asite DataFrame:\n{asite_df.to_string()}")
    logger.info(f"COS DataFrame:\n{cos_df.to_string()}")
    st.write("### Asite DataFrame (Debug):")
    st.write(asite_df)
    st.write("### COS DataFrame (Debug):")
    st.write(cos_df)

    combined_data = {
        "COS": cos_df,
        "Asite": asite_df
    }

    with st.spinner("Categorizing activities with WatsonX..."):
        ai_response = generatePrompt(combined_data, st.session_state.slabreport)
        st.session_state.ai_response = ai_response

    st.write("### Categorized Activity Counts (COS and Asite):")
    try:
        ai_data = json.loads(ai_response)
        st.json(ai_data)
    except json.JSONDecodeError as e:
        st.error(f"Failed to parse AI response as JSON: {str(e)}")
        st.write("Raw AI response:")
        st.text(ai_response)

    end_time = time.time()
    st.write(f"Total execution time: {end_time - start_time:.2f} seconds")
    
# COS File Fetching Function
@function_timer()
def get_cos_files():
    try:
        cos_client = initialize_cos_client()
        if not cos_client:
            st.error("❌ Failed to initialize COS client.")
            return []

        response = cos_client.list_objects_v2(Bucket=COS_BUCKET, Prefix="Veridia/")
        if 'Contents' not in response:
            st.error(f"❌ No files found in the 'Veridia' folder of bucket '{COS_BUCKET}'. Please ensure the folder exists and contains files.")
            logger.error("No objects found in Veridia folder")
            return []

        all_files = [obj['Key'] for obj in response.get('Contents', [])]
        st.write("**All files in Veridia folder:**")
        if all_files:
            st.write("\n".join(all_files))
        else:
            st.write("No files found.")
            logger.warning("Veridia folder is empty")
            return []

        # Pattern for Finishing Tracker files
        finishing_pattern = re.compile(
            r"Veridia/Tower\s*([4|5])\s*Finishing\s*Tracker[\(\s]*(.*?)(?:[\)\s]*\.xlsx)$",
            re.IGNORECASE
        )
        # Pattern for Anti. Slab Cycle file
        slab_cycle_pattern = re.compile(
            r"Veridia/Veridia Anti\. Slab Cycle With Possesion dates.*\.xlsx$",
            re.IGNORECASE
        )

        date_formats = [
            "%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"
        ]

        file_info = []
        for obj in response.get('Contents', []):
            key = obj['Key']
            # Check for Finishing Tracker files
            finishing_match = finishing_pattern.match(key)
            if finishing_match:
                tower_num = finishing_match.group(1)
                date_str = finishing_match.group(2).strip('()').strip()
                parsed_date = None
                
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.strptime(date_str, fmt)
                        break
                    except ValueError:
                        continue
                
                if parsed_date:
                    file_info.append({
                        'key': key,
                        'tower': tower_num,
                        'date': parsed_date,
                        'type': 'finishing'
                    })
                else:
                    logger.warning(f"Could not parse date in filename: {key} (date: {date_str})")
                    st.warning(f"Skipping file with unparseable date: {key}")
            # Check for Anti. Slab Cycle file
            elif slab_cycle_pattern.match(key):
                # No date parsing needed since the filename doesn't include a date
                file_info.append({
                    'key': key,
                    'tower': None,  # No specific tower associated
                    'date': obj['LastModified'],  # Use LastModified timestamp
                    'type': 'slab_cycle'
                })

        if not file_info:
            st.error("❌ No Excel files matched the expected patterns in the 'Veridia' folder. Expected formats: 'Tower 4/5 Finishing Tracker(date).xlsx' or 'Veridia Anti. Slab Cycle With Possesion dates*.xlsx'.")
            logger.error("No files matched the expected patterns")
            return []

        # Separate Finishing and Slab Cycle files
        finishing_files = {}
        slab_cycle_files = []
        for info in file_info:
            if info['type'] == 'finishing':
                tower = info['tower']
                if tower not in finishing_files or info['date'] > finishing_files[tower]['date']:
                    finishing_files[tower] = info
            elif info['type'] == 'slab_cycle':
                slab_cycle_files.append(info)

        # Select the latest Slab Cycle file (if multiple exist)
        if slab_cycle_files:
            latest_slab_file = max(slab_cycle_files, key=lambda x: x['date'])
            files = [info['key'] for info in finishing_files.values()] + [latest_slab_file['key']]
        else:
            files = [info['key'] for info in finishing_files.values()]

        if not files:
            st.error("❌ No valid Excel files found for Tower 4, Tower 5, or Anti. Slab Cycle after filtering.")
            logger.error("No valid files after filtering")
            return []

        st.success(f"Found {len(files)} matching files: {', '.join(files)}")
        return files
    except Exception as e:
        st.error(f"❌ Error fetching COS files: {str(e)}")
        logger.error(f"Error fetching COS files: {str(e)}")
        return []

# Initialize session state variables
if 'cos_df_tower4a' not in st.session_state:
    st.session_state.cos_df_tower4a = None
if 'cos_df_tower4b' not in st.session_state:
    st.session_state.cos_df_tower4b = None
if 'cos_df_tower5' not in st.session_state:
    st.session_state.cos_df_tower5 = None
if 'cos_tname_tower4a' not in st.session_state:
    st.session_state.cos_tname_tower4a = None
if 'cos_tname_tower4b' not in st.session_state:
    st.session_state.cos_tname_tower4b = None
if 'cos_tname_tower5' not in st.session_state:
    st.session_state.cos_tname_tower5 = None

# ADD THESE MISSING INITIALIZATIONS:
if 'cos_client' not in st.session_state:
    st.session_state.cos_client = None
if 'bucket_name' not in st.session_state:
    st.session_state.bucket_name = None
if 'file_list' not in st.session_state:
    st.session_state.file_list = None
if 'slabreport' not in st.session_state:
    st.session_state.slabreport = pd.DataFrame()
if 'slab_df' not in st.session_state:
    st.session_state.slab_df = pd.DataFrame()



if 'ignore_month' not in st.session_state:
    st.session_state.ignore_month = False  
if 'ignore_year' not in st.session_state:
    st.session_state.ignore_year = False 


# Process Excel files
@function_timer()
def process_file(file_stream, filename):
    try:
        workbook = openpyxl.load_workbook(file_stream)
        available_sheets = workbook.sheetnames
        logger.info(f"Available sheets in {filename}: {available_sheets}")

        # Check if the file is an Anti. Slab Cycle file
        is_slab_cycle = "Anti. Slab Cycle" in filename

        if is_slab_cycle:
            # Handle Anti. Slab Cycle file
            possible_sheet_names = [
                "Slab Cycle", "Anti Slab Cycle", "Veridia Slab Cycle",
                "Possession Dates", "SlabCycle", "AntiSlabCycle", "Annexure-01"
            ]
            sheet_name = None
            for name in possible_sheet_names:
                if name in available_sheets:
                    sheet_name = name
                    break
            if not sheet_name:
                # Fallback to the first sheet if no expected name is found
                sheet_name = available_sheets[0] if available_sheets else None
            if not sheet_name:
                st.error(f"No valid sheets found in {filename}. Available sheets: {', '.join(available_sheets)}")
                logger.error(f"No valid sheets found in {filename}")
                return (None, None), (None, None)

            file_stream.seek(0)
            try:
                # Try different header rows (0, 1, 2) to find valid columns
                df = None
                actual_columns = None
                for header_row in [0, 1, 2]:
                    try:
                        file_stream.seek(0)
                        df = pd.read_excel(file_stream, sheet_name=sheet_name, header=header_row)
                        actual_columns = df.columns.tolist()
                        # Check if columns are all "Unnamed"
                        if all(col.startswith("Unnamed:") for col in actual_columns):
                            continue
                        logger.info(f"Columns in {sheet_name} (header={header_row}): {actual_columns}")
                        st.write(f"Columns in {sheet_name} (header row {header_row + 1}): {actual_columns}")
                        st.write(f"First 5 rows of {sheet_name} (header row {header_row + 1}):")
                        st.write(df.head(5))
                        break
                    except Exception as e:
                        logger.warning(f"Failed to read {sheet_name} with header row {header_row}: {str(e)}")
                        continue

                if df is None or actual_columns is None:
                    st.error(f"Could not find valid headers in {sheet_name}. All attempts yielded unnamed columns or errors.")
                    logger.error(f"No valid headers found in {sheet_name}")
                    return (None, None), (None, None)

                # Define possible column names for mapping
                column_mapping = {
                    'Activity ID': ['Activity ID', 'Task ID', 'ID', 'ActivityID'],
                    'Activity Name': ['Activity Name', 'Task Name', 'Activity', 'Name', 'Task'],
                    'Actual Finish': ['Actual Finish', 'Finish Date', 'Completion Date', 'Actual End', 'End Date']
                }

                # Find matching columns
                target_columns = ['Activity ID', 'Activity Name', 'Actual Finish']
                selected_columns = {}
                for target in target_columns:
                    for possible_name in column_mapping[target]:
                        if possible_name in actual_columns:
                            selected_columns[target] = possible_name
                            break

                # Ensure critical columns are present
                if 'Activity Name' not in selected_columns or 'Actual Finish' not in selected_columns:
                    logger.error(f"Critical columns missing in {sheet_name}. Found: {selected_columns}")
                    return (None, None), (None, None)

                # Select and rename columns
                df = df[list(selected_columns.values())]
                df.columns = list(selected_columns.keys())
                
                df = df.dropna(subset=['Activity Name'])
                df['Activity Name'] = df['Activity Name'].astype(str).str.strip()

                if 'Actual Finish' in df.columns:
                    df['Actual_Finish_Original'] = df['Actual Finish'].astype(str)
                    df['Actual Finish'] = pd.to_datetime(df['Actual Finish'], errors='coerce')
                    has_na_mask = (
                        pd.isna(df['Actual Finish']) |
                        (df['Actual_Finish_Original'].str.upper() == 'NAT') |
                        (df['Actual_Finish_Original'].str.lower().isin(['nan', 'na', 'n/a', 'none', '']))
                    )
                    na_rows = df[has_na_mask][['Activity Name', 'Actual Finish']]
                    if not na_rows.empty:
                        st.write(f"Sample of rows with NA or invalid values in Actual Finish for {filename}:")
                        st.write(na_rows.head(10))
                        na_activities = na_rows.groupby('Activity Name').size().reset_index(name='Count')
                        st.write(f"Activities with NA or invalid Actual Finish values in {filename}:")
                        st.write(na_activities)
                    else:
                        st.write(f"No NA or invalid values found in Actual Finish for {filename}")
                    df.drop('Actual_Finish_Original', axis=1, inplace=True)

                st.write(f"Unique Activity Names in {sheet_name} ({filename}):")
                unique_activities = df[['Activity Name']].drop_duplicates()
                st.write(unique_activities)

                return (df, "Slab Cycle"), (None, None)
            except Exception as e:
                st.error(f"Error processing sheet {sheet_name} in {filename}: {str(e)}")
                logger.error(f"Error processing sheet {sheet_name} in {filename}: {str(e)}")
                return (None, None), (None, None)
        else:
            # Handle Tower 4 or Tower 5 Finishing Tracker files
            tower_num = None
            if "Tower 5" in filename or "Tower5" in filename:
                tower_num = "5"
            elif "Tower 4" in filename or "Tower4" in filename:
                tower_num = "4"

            if not tower_num:
                st.error(f"Cannot determine tower number from filename: {filename}")
                logger.error(f"Cannot determine tower number from filename: {filename}")
                return (None, None), (None, None)

            possible_sheet_names = [
                f"TOWER {tower_num} FINISHING",
                f"TOWER {tower_num} FINISHING.",
                f"TOWER{tower_num}FINISHING",
                f"TOWER{tower_num}FINISHING.",
                f"TOWER {tower_num}FINISHING",
                f"TOWER{tower_num} FINISHING",
                f"Tower {tower_num} Finishing",
                f"Finish"
            ]

            sheet_name = None
            for name in possible_sheet_names:
                if name in available_sheets:
                    sheet_name = name
                    break

            if not sheet_name:
                for available in available_sheets:
                    if f"TOWER {tower_num}" in available.upper() and "FINISH" in available.upper():
                        sheet_name = available
                        break

            if not sheet_name:
                st.error(f"Required sheet for Tower {tower_num} not found in file. Available sheets: {', '.join(available_sheets)}")
                logger.error(f"Required sheet for Tower {tower_num} not found in {filename}")
                return (None, None), (None, None)

            file_stream.seek(0)

            try:
                df = pd.read_excel(file_stream, sheet_name=sheet_name, header=0)

                expected_columns = [
                    'Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name',
                    'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish',
                    'Actual Start', 'Actual Finish', '% Complete', 'Start', 'Finish', 'Delay Reasons'
                ]

                if len(df.columns) < len(expected_columns):
                    st.warning(f"Excel file has fewer columns than expected ({len(df.columns)} found, {len(expected_columns)} expected).")
                    expected_columns = expected_columns[:len(df.columns)]

                df.columns = expected_columns[:len(df.columns)]

                target_columns = ["Module", "Floor", "Flat", "Activity ID", "Activity Name", "Actual Finish"]
                available_columns = [col for col in target_columns if col in df.columns]

                if len(available_columns) < len(target_columns):
                    missing_cols = [col for col in target_columns if col not in available_columns]
                    st.warning(f"Missing columns in file: {', '.join(missing_cols)}")
                    for col in missing_cols:
                        df[col] = None

                df = df[target_columns]
                df = df.dropna(subset=['Activity Name'])

                df['Activity Name'] = df['Activity Name'].astype(str).str.strip()

                if 'Floor' in df.columns:
                    df['Floor'] = df['Floor'].astype(str)
                    v_rows = df[df['Floor'].str.strip().str.upper() == 'V']
                    if not v_rows.empty:
                        df = pd.concat([df, v_rows], ignore_index=True)

                if 'Actual Finish' in df.columns:
                    df['Actual_Finish_Original'] = df['Actual Finish'].astype(str)
                    df['Actual Finish'] = pd.to_datetime(df['Actual Finish'], errors='coerce')
                    has_na_mask = (
                        pd.isna(df['Actual Finish']) |
                        (df['Actual_Finish_Original'].str.upper() == 'NAT') |
                        (df['Actual_Finish_Original'].str.lower().isin(['nan', 'na', 'n/a', 'none', '']))
                    )
                    na_rows = df[has_na_mask][['Activity Name', 'Actual Finish']]
                    if not na_rows.empty:
                        st.write("Sample of rows with NA or invalid values in Actual Finish:")
                        st.write(na_rows.head(10))
                        na_activities = na_rows.groupby('Activity Name').size().reset_index(name='Count')
                        st.write("Activities with NA or invalid Actual Finish values:")
                        st.write(na_activities)
                    else:
                        st.write("No NA or invalid values found in Actual Finish")
                    df.drop('Actual_Finish_Original', axis=1, inplace=True)

                st.write(f"Unique Activity Names in {sheet_name}:")
                unique_activities = df[['Module', 'Floor', 'Activity Name']].drop_duplicates()
                st.write(unique_activities)

                if tower_num == "4":
                    df['Module'] = df['Module'].astype(str).str.strip().str.upper()
                    modules_a = ['M5', 'M6', 'M7', 'M8']
                    modules_b = ['M1', 'M2', 'M3', 'M4']
                    mask_a = df['Module'].isin(modules_a)
                    mask_b = df['Module'].isin(modules_b)
                    df_tower4a = df[mask_a].copy()
                    df_tower4b = df[mask_b].copy()
                    st.write(f"Tower 4(A) (Modules M5-M8) - {len(df_tower4a)} rows:")
                    st.write(df_tower4a.head())
                    st.write(f"Tower 4(B) (Modules M1-M4) - {len(df_tower4b)} rows:")
                    st.write(df_tower4b.head())
                    return (df_tower4a, "Tower 4(A)"), (df_tower4b, "Tower 4(B)")
                else:
                    return (df, f"Tower {tower_num}"), (None, None)

            except Exception as e:
                st.error(f"Error processing sheet {sheet_name}: {str(e)}")
                logger.error(f"Error processing sheet {sheet_name}: {str(e)}")
                return (None, None), (None, None)

    except Exception as e:
        st.error(f"Error loading Excel file {filename}: {str(e)}")
        logger.error(f"Error loading Excel file {filename}: {str(e)}")
        return (None, None), (None, None)
    
    
#Slab code
@function_timer()
def GetSlabReport():
    foundverdia = False
    today = date.today()
    prev_month = today - relativedelta(months=1)
    month_year = today.strftime("%m-%Y")
    prev_month_year = prev_month.strftime("%m-%Y")
    
    # cos_client = initialize_cos_client()
     
    try:
        logger.info("Attempting to initialize COS client...")
        cos_client = ibm_boto3.client(
            's3',
            ibm_api_key_id=COS_API_KEY,
            ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
            config=Config(
                signature_version='oauth',
                connect_timeout=180,
                read_timeout=180,
                retries={'max_attempts': 15}
            ),
            endpoint_url=COS_ENDPOINT
        )
        response = cos_client.list_objects_v2(Bucket="projectreportnew")
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]

        for file in files:
            
            try:
                if file.startswith("Veridia") and "Structure Work Tracker" in file:
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    
                    if st.session_state.ignore_month and st.session_state.ignore_year:
                        st.session_state.slabreport = ProcessVeridia(io.BytesIO(response['Body'].read()), st.session_state.ignore_year, st.session_state.ignore_month)
                    else:
                        st.session_state.slabreport = ProcessVeridia(io.BytesIO(response['Body'].read()))              
                    foundverdia = True
                   
                    break
                    
            except Exception as e:
                st.info(e)
                st.session_state.slabreport = "No Data Found"

        if not foundverdia:
            for file in files:
                try:
                    if file.startswith("Veridia") and "Structure Work Tracker" in file:
                        # st.write("🕓 Previous month:", file)
                        response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                        # st.session_state.slabreport = ProcessVeridia(io.BytesIO(response['Body'].read()))
                        if st.session_state.ignore_month and st.session_state.ignore_year:
                            st.session_state.slabreport = ProcessVeridia(io.BytesIO(response['Body'].read()), st.session_state.ignore_year, st.session_state.ignore_month)              # st.write(veridia)
                        
                        break
                    # return veridia
                except Exception as e:
                    st.error(e)
                    st.session_state.slabreport = "No Data Found"
                   
    except Exception as e:
        print(f"Error fetching COS files: {e}")
        files = ["Error fetching COS files"]
        st.session_state.slabreport = "No Data Found"


@function_timer()
def generatePrompt(combined_data, slab):
    try:
        st.write(slab)
        st.write(json.loads(slab))
        cos_df = combined_data["COS"] if isinstance(combined_data["COS"], pd.DataFrame) else pd.DataFrame()
        asite_df = combined_data["Asite"] if isinstance(combined_data["Asite"], pd.DataFrame) else pd.DataFrame()

        cos_json = cos_df[['Tower', 'Activity Name', 'Count']].to_json(orient='records', indent=2)
        asite_json = asite_df[['Tower', 'Activity Name', 'Count']].to_json(orient='records', indent=2)

        body = {
            "input": f"""
            Read the table data provided below for COS and Asite sources, which include tower-specific activity counts. Categorize the activities into the specified categories (MEP, Interior Finishing, Structure Work, ED Civil) for each tower in each source (COS and Asite). Compute the total count of each activity within its respective category for each tower and return the results as a JSON object with "COS" and "Asite" sections, where each section contains a list of towers, each with categories and their activities. For the MEP category in COS, calculate the total count between 'UP-First Fix' and 'CP-First Fix' and report it as 'Min. count of UP-First Fix and CP-First Fix' for each tower. If an activity is not found for a tower, include it with a count of 0. If the Structure Work category has no activities in COS, return an empty list for it. Ensure the counts are accurate, the output is grouped by tower and category, and the JSON structure is valid with no nested or repeated keys.

            The data provided is as follows:
            

            Slab:
            {slab}

            COS Table Data:
            {cos_json}

            Asite Table Data:
            {asite_json}

            Categories and Activities:
            COS:
            - MEP: EL-First Fix, UP-First Fix, CP-First Fix, Min. count of UP-First Fix and CP-First Fix, C-Gypsum and POP Punning, EL-Second Fix, Concreting, Electrical
            - Interior Finishing: Installation of doors, Waterproofing Works, Wall Tiling, Floor Tiling
            - ED Civil: Sewer Line, Storm Line, GSB, WMM, Stamp Concrete, Saucer drain, Kerb Stone
            Asite:
            - MEP: Wall Conducting, Plumbing Works, POP & Gypsum Plaster, Wiring & Switch Socket, Slab Conducting, Electrical Cable, Concreting
            - Interior Finishing: Door/Window Frame, Waterproofing - Sunken, Wall Tiling, Floor Tiling, Door/Window Shutter
            - ED Civil: Sewer Line, Rain Water/Storm Line, Granular Sub-base, WMM, Saucer drain/Paver block, Kerb Stone, Concreting

            Slab:
            - Get total greens of Each Tower

            Example JSON format needed:
            {{
              "COS": [
                {{
                  "Tower": "Tower 4(A)",
                  "Categories": [
                    {{
                      "Category": "MEP",
                      "Activities": [
                        {{"Activity Name": "EL-First Fix", "Total": 0}},
                        {{"Activity Name": "UP-First Fix", "Total": 0}},
                        {{"Activity Name": "CP-First Fix", "Total": 0}},
                        {{"Activity Name": "Min. count of UP-First Fix and CP-First Fix", "Total": 0}},
                        {{"Activity Name": "C-Gypsum and POP Punning", "Total": 0}},
                        {{"Activity Name": "EL-Second Fix", "Total": 0}},
                        {{"Activity Name": "Concreting", "Total": 0}},
                        {{"Activity Name": "Electrical", "Total": 0}}
                      ]
                    }},
                    {{
                      "Category": "Interior Finishing",
                      "Activities": [
                        {{"Activity Name": "Installation of doors", "Total": 0}},
                        {{"Activity Name": "Waterproofing Works", "Total": 0}},
                        {{"Activity Name": "Wall Tiling", "Total": 0}},
                        {{"Activity Name": "Floor Tiling", "Total": 0}}
                      ]
                    }},
                    {{
                      "Category": "ED Civil",
                      "Activities": [
                        {{"Activity Name": "Sewer Line", "Total": 0}},
                        {{"Activity Name": "Storm Line", "Total": 0}},
                        {{"Activity Name": "GSB", "Total": 0}},
                        {{"Activity Name": "WMM", "Total": 0}},
                        {{"Activity Name": "Stamp Concrete", "Total": 0}},
                        {{"Activity Name": "Saucer drain", "Total": 0}},
                        {{"Activity Name": "Kerb Stone", "Total": 0}}
                      ]
                    }}
                  ]
                }},
                {{ "Tower": "Tower 4(B)", "Categories": [...] }},
                {{ "Tower": "Tower 5", "Categories": [...] }}
              ],
              "Asite": [
                {{
                  "Tower": "Tower 4(A)",
                  "Categories": [
                    {{
                      "Category": "MEP",
                      "Activities": [
                        {{"Activity Name": "Wall Conducting", "Total": 0}},
                        {{"Activity Name": "Plumbing Works", "Total": 0}},
                        {{"Activity Name": "POP & Gypsum Plaster", "Total": 0}},
                        {{"Activity Name": "Wiring & Switch Socket", "Total": 0}},
                        {{"Activity Name": "Slab Conducting", "Total": 0}},
                        {{"Activity Name": "Electrical Cable", "Total": 0}},
                        {{"Activity Name": "Concreting", "Total": 0}}
                      ]
                    }},
                    {{
                      "Category": "Interior Finishing",
                      "Activities": [
                        {{"Activity Name": "Door/Window Frame", "Total": 0}},
                        {{"Activity Name": "Waterproofing - Sunken", "Total": 0}},
                        {{"Activity Name": "Wall Tiling", "Total": 0}},
                        {{"Activity Name": "Floor Tiling", "Total": 0}},
                        {{"Activity Name": "Door/Window Shutter", "Total": 0}}
                      ]
                    }},
                    {{
                      "Category": "ED Civil",
                      "Activities": [
                        {{"Activity Name": "Sewer Line", "Total": 0}},
                        {{"Activity Name": "Rain Water/Storm Line", "Total": 0}},
                        {{"Activity Name": "Granular Sub-base", "Total": 0}},
                        {{"Activity Name": "WMM", "Total": 0}},
                        {{"Activity Name": "Saucer drain/Paver block", "Total": 0}},
                        {{"Activity Name": "Kerb Stone", "Total": 0}}
                      ]
                    }}
                  ]
                }},
                {{ "Tower": "Tower 4(B)", "Categories": [...] }},
                {{ "Tower": "Tower 5", "Categories": [...] }}
              ],
              "Slab":{{
                 "Tower Name":"Total"
              }}
            }}

            Return only the JSON object, no additional text, explanations, or code. Ensure the counts are accurate, activities are correctly categorized, and the JSON structure is valid.
            """,
            "parameters": {
                "decoding_method": "greedy",
                "max_new_tokens": 8100,
                "min_new_tokens": 0,
                "stop_sequences": [],  # Removed "}" as it can cause truncation issues
                "repetition_penalty": 1.0,
                "temperature": 0.1
            },
            "model_id": os.getenv("MODEL_ID_1"),
            "project_id": os.getenv("PROJECT_ID_1")
        }
        
        access_token = get_access_token(os.getenv("API_KEY_1"))
        if not access_token:
            logger.error("Failed to obtain access token for WatsonX API")
            return (combined_data)
            
        headers = {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Authorization": f"Bearer {access_token}"
        }
        
        response = requests.post(os.getenv("WATSONX_API_URL_1"), headers=headers, json=body, timeout=1000)
        
        if response.status_code != 200:
            logger.error(f"WatsonX API call failed: {response.status_code} - {response.text}")
            st.warning(f"WatsonX API failed with status {response.status_code}: {response.text}. Using fallback method to calculate totals.")
            return (combined_data)
            
        response_data = response.json()
        if 'results' not in response_data or not response_data['results']:
            logger.error("WatsonX API response does not contain 'results' key")
            st.warning("WatsonX API response invalid. Using fallback method to calculate totals.")
            return (combined_data)

        generated_text = response_data['results'][0].get('generated_text', '').strip()
        logger.info(f"Raw WatsonX API response: {generated_text[:1000]}...")
        if not generated_text:
            logger.error("WatsonX API returned empty generated text")
            st.warning("WatsonX API returned empty response. Using fallback method to calculate totals.")
            return (combined_data)

        # Fix 1: Enhanced JSON extraction with repair capability
        fixed_json_text = extract_and_repair_json(generated_text)
        if fixed_json_text is None:
            logger.error("Failed to extract or repair JSON from response")
            return (combined_data)
        
        try:
            parsed_json = json.loads(fixed_json_text)
            if not (isinstance(parsed_json, dict) and "COS" in parsed_json and "Asite" in parsed_json):
                logger.warning(f"Invalid JSON structure: {json.dumps(parsed_json, indent=2)}")
                return (combined_data)
            for source in ["COS", "Asite"]:
                if not isinstance(parsed_json[source], list):
                    logger.warning(f"Expected list for {source}, got: {type(parsed_json[source])}")
                    return (combined_data)
                for tower_data in parsed_json[source]:
                    if not isinstance(tower_data, dict) or "Tower" not in tower_data or "Categories" not in tower_data:
                        logger.warning(f"Invalid tower data in {source}: {tower_data}")
                        return (combined_data)
            return json.dumps(parsed_json, indent=2)  # Return standardized JSON
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON after repair attempt: {str(e)}")
            logger.error(f"Full response after repair: {fixed_json_text}")
            error_position = int(str(e).split('(char ')[1].split(')')[0]) if '(char ' in str(e) else 0
            context_start = max(0, error_position - 50)
            context_end = min(len(fixed_json_text), error_position + 50)
            logger.error(f"JSON error context: ...{fixed_json_text[context_start:error_position]}[ERROR HERE]{fixed_json_text[error_position:context_end]}...")
            st.warning(f"WatsonX API returned invalid JSON that couldn't be repaired. Error: {str(e)}. Using fallback method to calculate totals.")
            return (combined_data)
    
    except Exception as e:
        logger.error(f"Error in WatsonX API call: {str(e)}")
        st.warning(f"Error in WatsonX API call: {str(e)}. Using fallback method to calculate totals.")
        return (combined_data)


@function_timer()
def extract_and_repair_json(text):
    # Try to find JSON content within the response
    json_match = re.search(r'\{.*\}', text, re.DOTALL)
    if json_match:
        json_text = json_match.group(0)
        
        # Common JSON repair operations
        try:
            # First try: Parse as is
            json.loads(json_text)
            return json_text
        except json.JSONDecodeError as e:
            try:
                # Fix 1: Try fixing missing commas between objects in arrays
                fixed1 = re.sub(r'}\s*{', '},{', json_text)
                json.loads(fixed1)
                logger.info("JSON fixed by adding missing commas between objects")
                return fixed1
            except json.JSONDecodeError:
                try:
                    # Fix 2: Try fixing trailing commas in arrays/objects
                    fixed2 = re.sub(r',\s*}', '}', fixed1)
                    fixed2 = re.sub(r',\s*]', ']', fixed2)
                    json.loads(fixed2)
                    logger.info("JSON fixed by removing trailing commas")
                    return fixed2
                except json.JSONDecodeError:
                    try:
                        # Fix 3: Try using a JSON repair library or more aggressive repairs
                        # Here we'll use a simple approach to balance braces/brackets
                        fixed3 = fixed2
                        count_open_braces = fixed3.count('{')
                        count_close_braces = fixed3.count('}')
                        if count_open_braces > count_close_braces:
                            fixed3 += '}' * (count_open_braces - count_close_braces)
                        
                        count_open_brackets = fixed3.count('[')
                        count_close_brackets = fixed3.count(']')
                        if count_open_brackets > count_close_brackets:
                            fixed3 += ']' * (count_open_brackets - count_close_brackets)
                        
                        # Fix unquoted keys (a common issue)
                        fixed3 = re.sub(r'([{,])\s*([a-zA-Z0-9_]+)\s*:', r'\1"\2":', fixed3)
                        
                        json.loads(fixed3)
                        logger.info("JSON fixed with aggressive repairs")
                        return fixed3
                    except json.JSONDecodeError:
                        # Final attempt: Try to load the JSON with a more permissive parser
                        try:
                            import demjson3  # type: ignore
                            parsed = demjson3.decode(json_text)
                            logger.info("JSON fixed with demjson3")
                            return json.dumps(parsed)
                        except Exception:
                            logger.error("All JSON repair attempts failed")
                            return None
    else:
        logger.error("No JSON-like content found in the response")
        return None

# Fix the getTotal function to handle the improved JSON structure
@function_timer()
def getTotal(ai_data):
    try:
        if isinstance(ai_data, str):
            try:
                ai_data = json.loads(ai_data)
            except json.JSONDecodeError as e:
                logger.error(f"Error parsing AI data JSON: {str(e)}")
                return [0] * len(st.session_state.get('sheduledf', pd.DataFrame()).index)
            
        if not isinstance(ai_data, dict) or "COS" not in ai_data or "Asite" not in ai_data:
            logger.error(f"AI data is not in expected format: {ai_data}")
            return [0] * len(st.session_state.get('sheduledf', pd.DataFrame()).index)

        share = []
        
        # Process sources properly
        for source in ["COS", "Asite"]:
            if not isinstance(ai_data[source], list):
                logger.error(f"{source} data is not a list: {ai_data[source]}")
                continue
                
            for tower_data in ai_data[source]:
                if not isinstance(tower_data, dict) or "Categories" not in tower_data:
                    logger.error(f"Invalid tower data format in {source}: {tower_data}")
                    continue
                    
                for category_data in tower_data["Categories"]:
                    if not isinstance(category_data, dict) or "Activities" not in category_data:
                        logger.error(f"Invalid category data format: {category_data}")
                        continue
                        
                    for activity in category_data["Activities"]:
                        if isinstance(activity, dict) and "Total" in activity:
                            total = activity["Total"]
                            share.append(int(total) if isinstance(total, (int, float)) and pd.notna(total) else 0)
                        else:
                            logger.warning(f"Activity missing Total field: {activity}")
                            share.append(0)
        
        # Ensure we have enough values for the schedule dataframe
        expected_length = len(st.session_state.get('sheduledf', pd.DataFrame()).index)
        if len(share) < expected_length:
            logger.warning(f"Not enough values in share list (got {len(share)}, need {expected_length}). Padding with zeros.")
            share.extend([0] * (expected_length - len(share)))
        elif len(share) > expected_length:
            logger.warning(f"Too many values in share list (got {len(share)}, need {expected_length}). Truncating.")
            share = share[:expected_length]
            
        return share
    except Exception as e:
        logger.error(f"Error processing AI data: {str(e)}")
        st.error(f"Error processing AI data: {str(e)}")
        return [0] * len(st.session_state.get('sheduledf', pd.DataFrame()).index)    
 

   
# Function to handle activity count display logic
@function_timer()
def display_activity_count():
    try:
        if 'ai_response' not in st.session_state or not st.session_state.ai_response:
            st.error("❌ No AI-generated data available. Please run the analysis first.")
            return

        try:
            ai_data = json.loads(st.session_state.ai_response)
        except json.JSONDecodeError as e:
            st.error(f"❌ Failed to parse AI response: {str(e)}")
            st.write("Raw AI response:")
            st.text(st.session_state.ai_response)
            return

        if not isinstance(ai_data, dict) or "COS" not in ai_data or "Asite" not in ai_data:
            st.error("❌ Invalid AI data format. Expected 'COS' and 'Asite' sections.")
            st.write("AI data content:")
            st.json(ai_data)
            return

        slab_df = st.session_state.get('slab_df', pd.DataFrame())
        logging.info(f"Slab cycle DataFrame in display_activity_count: {slab_df.to_dict()}")
        slab_display_df = pd.DataFrame(columns=['Tower', 'Completed'])
        slab_counts = {}
        if not slab_df.empty:
            new_rows = []
            for _, row in slab_df.iterrows():
                tower = row['Tower']
                completed = row['Completed']
                if tower == 'T4':
                    t4a_completed = completed // 2
                    t4b_completed = completed - t4a_completed
                    new_rows.append({'Tower': 'T4A', 'Completed': t4a_completed})
                    new_rows.append({'Tower': 'T4B', 'Completed': t4b_completed})
                    slab_counts['T4A'] = t4a_completed
                    slab_counts['T4B'] = t4b_completed
                else:
                    new_rows.append({'Tower': tower, 'Completed': completed})
                    slab_counts[tower] = completed
            slab_display_df = pd.DataFrame(new_rows)
            logging.info(f"Slab display DataFrame after processing: {slab_display_df.to_dict()}")
        else:
            logging.warning("Slab cycle DataFrame is empty in display_activity_count.")

        categories = {
            "COS": {
                "MEP": [
                    "EL-First Fix", "UP-First Fix", "CP-First Fix", "Min. count of UP-First Fix and CP-First Fix",
                    "C-Gypsum and POP Punning", "EL-Second Fix", "Concreting", "Electrical"
                ],
                "Interior Finishing": [
                    "Installation of doors", "Waterproofing Works", "Wall Tiling", "Floor Tiling"
                ],
                "ED Civil": [
                    "Sewer Line", "Storm Line", "GSB", "WMM", "Stamp Concrete", "Saucer drain", "Kerb Stone"
                ]
            },
            "Asite": {
                "MEP": [
                    "Wall Conducting", "Plumbing Works", "POP & Gypsum Plaster", "Wiring & Switch Socket",
                    "Slab Conducting", "Electrical Cable", "Concreting"
                ],
                "Interior Finishing": [
                    "Door/Window Frame", "Waterproofing - Sunken", "Wall Tile", "Floor Tile", "Door/Window Shutter"
                ],
                "ED Civil": [
                    "Sewer Line", "Rain Water/Storm Line", "Granular Sub-base", "WMM",
                    "Saucer drain/Paver block", "Kerb Stone", "Concreting"
                ]
            }
        }

        for source in ["COS", "Asite"]:
            st.subheader(f"{source} Activity Counts")
            source_data = ai_data.get(source, [])
            if not source_data:
                st.warning(f"No data available for {source}.")
                continue

            for tower_data in source_data:
                tower_name = tower_data.get("Tower", "Unknown Tower")
                st.write(f"#### {tower_name}")
                tower_categories = tower_data.get("Categories", [])

                if not tower_categories:
                    st.write("No categories available for this tower.")
                    continue

                tower_total = 0

                for category in categories[source]:
                    st.write(f"**{category}**")
                    category_data = next(
                        (cat for cat in tower_categories if cat.get("Category") == category),
                        {"Category": category, "Activities": []}
                    )

                    if not category_data["Activities"]:
                        st.write("No activities recorded.")
                        continue

                    activity_counts = []
                    for activity in categories[source][category]:
                        activity_info = next(
                            (act for act in category_data["Activities"] if act.get("Activity Name") == activity),
                            {"Activity Name": activity, "Total": 0}
                        )
                        count = int(activity_info["Total"]) if pd.notna(activity_info["Total"]) else 0
                        activity_counts.append({
                            "Activity Name": activity_info["Activity Name"],
                            "Count": count
                        })
                        tower_total += count

                    df = pd.DataFrame(activity_counts)
                    if not df.empty:
                        st.table(df)
                    else:
                        st.write("No activities in this category.")

                if source == "COS":
                    st.write("**Slab Cycle Counts**")
                    tower_slab_df = slab_display_df[slab_display_df['Tower'] == tower_name]
                    logging.info(f"Tower {tower_name} - Filtered slab counts: {tower_slab_df.to_dict()}")
                    if not tower_slab_df.empty:
                        st.table(tower_slab_df)
                        tower_total += tower_slab_df['Completed'].sum()
                    else:
                        st.write("No slab cycle data for this tower.")
                        st.write("All available slab cycle counts (debug):")
                        st.table(slab_display_df)

                st.write(f"**Total for {tower_name}**: {tower_total}")

        total_cos = sum(
            act["Total"]
            for tower in ai_data.get("COS", [])
            for cat in tower.get("Categories", [])
            for act in cat.get("Activities", [])
            if isinstance(act.get("Total", 0), (int, float)) and pd.notna(act["Total"])
        )
        total_cos += sum(slab_counts.values())

        total_asite = sum(
            act["Total"]
            for tower in ai_data.get("Asite", [])
            for cat in tower.get("Categories", [])
            for act in cat.get("Activities", [])
            if isinstance(act.get("Total", 0), (int, float)) and pd.notna(act["Total"])
        )

        st.write("### Total Completed Activities")
        st.write(f"**COS Total**: {total_cos}")
        st.write(f"**Asite Total**: {total_asite}")

    except Exception as e:
        logging.error(f"Error in display_activity_count: {str(e)}")
        st.error(f"❌ Error displaying activity counts: {str(e)}")
        st.write("AI response content (for debugging):")
        st.text(st.session_state.ai_response)
        
st.markdown(
    """
    <h1 style='font-family: "Arial Black", Gadget, sans-serif; 
               color: red; 
               font-size: 48px; 
               text-align: center;'>
        CheckList - Report
    </h1>
    """,
    unsafe_allow_html=True
)



# Combined function for Initialize All Data and Fetch COS
@function_timer()
async def initialize_and_fetch_data(email, password):
    with st.spinner("Starting initialization and data fetching process..."):
        # Step 1: Login
        if not email or not password:
            st.sidebar.error("Please provide both email and password!")
            return False
        try:
            st.sidebar.write("Logging in...")
            session_id = await login_to_asite(email, password)
            if not session_id:
                st.sidebar.error("Login failed!")
                return False
            st.sidebar.success("Login successful!")
        except Exception as e:
            st.sidebar.error(f"Login failed: {str(e)}")
            return False

        # Step 2: Get Workspace ID
        try:
            st.sidebar.write("Fetching Workspace ID...")
            await GetWorkspaceID()
            st.sidebar.success("Workspace ID fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Workspace ID: {str(e)}")
            return False

        # Step 3: Get Project IDs
        try:
            st.sidebar.write("Fetching Project IDs...")
            await GetProjectId()
            st.sidebar.success("Project IDs fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Project IDs: {str(e)}")
            return False

        # Step 4: Get All Data
        try:
            st.sidebar.write("Fetching All Data...")
            veridiafinishing, veridiastructure, veridiaexternal, veridialift, veridiacommonarea = await GetAllDatas()
            st.session_state.veridiafinishing = veridiafinishing
            st.session_state.veridiastructure = veridiastructure
            st.session_state.veridiaexternal = veridiaexternal  
            st.session_state.veridialift = veridialift
            st.session_state.veridiacommonarea = veridiacommonarea
            st.sidebar.success("All Data fetched successfully!")
            logger.info(f"Stored veridiafinishing: {len(veridiafinishing)} records, veridiastructure: {len(veridiastructure)} records, veridiaexternal: {len(veridiaexternal)} records, veridialift: {len(veridialift)} records, veridia_common_area: {len(veridiacommonarea)} records")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch All Data: {str(e)}")
            logger.error(f"Failed to fetch All Data: {str(e)}")
            return False

       
        # Step 5: Get Activity Data
        try:
            st.sidebar.write("Fetching Activity Data...")
            finishing_activity_data, structure_activity_data, external_activity_data, lift_activity_data, common_area_activity_data = await Get_Activity()
            # Validate DataFrames
            activity_dataframes = {
                "finishing_activity_data": finishing_activity_data,
                "structure_activity_data": structure_activity_data,
                "external_activity_data": external_activity_data,
                "lift_activity_data": lift_activity_data,
                "common_area_activity_data": common_area_activity_data
            }
            for name, df in activity_dataframes.items():
                if df is None:
                    logger.error(f"{name} is None")
                    raise ValueError(f"{name} is None")
                if not isinstance(df, pd.DataFrame):
                    logger.error(f"{name} is not a DataFrame: {type(df)}")
                    raise ValueError(f"{name} is not a valid DataFrame")
                logger.info(f"{name} has {len(df)} records, empty: {df.empty}")
                if df.empty:
                    logger.warning(f"{name} is empty")
            # Store in session state
            st.session_state.finishing_activity_data = finishing_activity_data
            st.session_state.structure_activity_data = structure_activity_data
            st.session_state.external_activity_data = external_activity_data
            st.session_state.lift_activity_data = lift_activity_data
            st.session_state.common_area_activity_data = common_area_activity_data
            st.sidebar.success("Activity Data fetched successfully!")
            logger.info(f"Stored activity data - Finishing: {len(finishing_activity_data)} records, "
                        f"Structure: {len(structure_activity_data)} records, "
                        f"External: {len(external_activity_data)} records, "
                        f"Lift: {len(lift_activity_data)} records, "
                        f"Common Area: {len(common_area_activity_data)} records")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Activity Data: {str(e)}")
            logger.error(f"Failed to fetch Activity Data: {str(e)}\nStack trace:\n{traceback.format_exc()}")
            return False

        # Step 6: Get Location/Module Data
        try:
            st.sidebar.write("Fetching Location/Module Data...")
            finishing_location_data, structure_location_data, external_location_data, lift_location_data, common_area_location_data = await Get_Location()
            # Validate DataFrames
            location_dataframes = {
                "finishing_location_data": finishing_location_data,
                "structure_location_data": structure_location_data,
                "external_location_data": external_location_data,
                "lift_location_data": lift_location_data,
                "common_area_location_data": common_area_location_data
            }
            for name, df in location_dataframes.items():
                if df is None:
                    logger.error(f"{name} is None")
                    raise ValueError(f"{name} is None")
                if not isinstance(df, pd.DataFrame):
                    logger.error(f"{name} is not a DataFrame: {type(df)}")
                    raise ValueError(f"{name} is not a valid DataFrame")
                logger.info(f"{name} has {len(df)} records, empty: {df.empty}")
                if df.empty:
                    logger.warning(f"{name} is empty")
            # Store in session state
            st.session_state.finishing_location_data = finishing_location_data
            st.session_state.structure_location_data = structure_location_data
            st.session_state.external_location_data = external_location_data
            st.session_state.lift_location_data = lift_location_data
            st.session_state.common_area_location_data = common_area_location_data
            st.sidebar.success("Location/Module Data fetched successfully!")
            logger.info(f"Stored location data - Finishing: {len(finishing_location_data)} records, "
                        f"Structure: {len(structure_location_data)} records, "
                        f"External: {len(external_location_data)} records, "
                        f"Lift: {len(lift_location_data)} records, "
                        f"Common Area: {len(common_area_location_data)} records")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Location/Module Data: {str(e)}")
            logger.error(f"Failed to fetch Location/Module Data: {str(e)}\nStack trace:\n{traceback.format_exc()}")
            return False
        
        # Step 7: Fetch COS Files
        try:
            st.sidebar.write("Fetching COS files from Veridia folder...")
            files = get_cos_files()
            st.session_state.files = files
            if files:
                st.success(f"Found {len(files)} files in COS storage")
                for selected_file in files:
                    try:
                        st.write(f"Processing file: {selected_file}")
                        cos_client = initialize_cos_client()
                        if not cos_client:
                            st.error("Failed to initialize COS client")
                            continue
                        response = cos_client.get_object(Bucket=COS_BUCKET, Key=selected_file)
                        file_bytes = io.BytesIO(response['Body'].read())
                        result = process_file(file_bytes, selected_file)
                        if len(result) == 2:  # Handle Tower 4 split
                            (df_first, tname_first), (df_second, tname_second) = result
                            if df_first is not None and not df_first.empty:
                                if "Tower 4(A)" in tname_first:
                                    st.session_state.cos_df_tower4a = df_first
                                    st.session_state.cos_tname_tower4a = tname_first
                                    st.write(f"Processed Data for {tname_first} - {len(df_first)} rows:")
                                    st.write(df_first.head())
                                elif "Tower 4(B)" in tname_first:
                                    st.session_state.cos_df_tower4b = df_first
                                    st.session_state.cos_tname_tower4b = tname_first
                                    st.write(f"Processed Data for {tname_first} - {len(df_first)} rows:")
                                    st.write(df_first.head())
                                elif "Tower 5" in tname_first:
                                    st.session_state.cos_df_tower5 = df_first
                                    st.session_state.cos_tname_tower5 = tname_first
                                    st.write(f"Processed Data for {tname_first} - {len(df_first)} rows:")
                                    st.write(df_first.head())
                            if df_second is not None and not df_second.empty:
                                if "Tower 4(A)" in tname_second:
                                    st.session_state.cos_df_tower4a = df_second
                                    st.session_state.cos_tname_tower4a = tname_second
                                    st.write(f"Processed Data for {tname_second} - {len(df_second)} rows:")
                                    st.write(df_second.head())
                                elif "Tower 4(B)" in tname_second:
                                    st.session_state.cos_df_tower4b = df_second
                                    st.session_state.cos_tname_tower4b = tname_second
                                    st.write(f"Processed Data for {tname_second} - {len(df_second)} rows:")
                                    st.write(df_second.head())
                        elif len(result) == 1:  # Handle Tower 5
                            (df_first, tname_first) = result[0]
                            if df_first is not None and not df_first.empty:
                                if "Tower 4(A)" in tname_first:
                                    st.session_state.cos_df_tower4a = df_first
                                    st.session_state.cos_tname_tower4a = tname_first
                                    st.write(f"Processed Data for {tname_first} - {len(df_first)} rows:")
                                    st.write(df_first.head())
                                elif "Tower 4(B)" in tname_first:
                                    st.session_state.cos_df_tower4b = df_first
                                    st.session_state.cos_tname_tower4b = tname_first
                                    st.write(f"Processed Data for {tname_first} - {len(df_first)} rows:")
                                    st.write(df_first.head())
                                elif "Tower 5" in tname_first:
                                    st.session_state.cos_df_tower5 = df_first
                                    st.session_state.cos_tname_tower5 = tname_first
                                    st.write(f"Processed Data for {tname_first} - {len(df_first)} rows:")
                                    st.write(df_first.head())
                            if "Tower 5" in selected_file:
                                st.info(f"Processed Tower 5 data successfully.")
                            else:
                                st.warning(f"No secondary data (Tower 4 split) for {selected_file}.")
                        else:
                            st.warning(f"No data found in {selected_file}.")
                    except Exception as e:
                        st.error(f"Error loading {selected_file} from cloud storage: {str(e)}")
                        logger.error(f"Error loading {selected_file}: {str(e)}")
            else:
                st.warning("No expected Excel files available in the 'Veridia' folder of the COS bucket.")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch COS files: {str(e)}")
            logger.error(f"Failed to fetch COS files: {str(e)}")
            return False

        # Step 8: Verify stored session state keys
        st.sidebar.write("Verifying stored data...")
        required_keys = [
            'veridiafinishing', 'veridiastructure', 'veridiaexternal',
            'finishing_activity_data', 'structure_activity_data', 'external_activity_data',
            'finishing_location_data', 'structure_location_data', 'external_location_data'
        ]
        missing_keys = [key for key in required_keys if key not in st.session_state]
        if missing_keys:
            st.sidebar.error(f"Missing session state keys: {', '.join(missing_keys)}")
            logger.error(f"Missing session state keys: {', '.join(missing_keys)}")
            return False
        else:
            st.sidebar.success("All required data stored successfully!")
            logger.info("All required session state keys verified.")

        st.sidebar.write("Initialization and data fetching process completed!")
        return True



@function_timer()
def generate_consolidated_Checklist_excel(ai_data):
    try:
        # Parse AI data if it's a string
        if isinstance(ai_data, str):
            logger.info("Parsing ai_data from string")
            ai_data = json.loads(ai_data)
        
        if not isinstance(ai_data, dict) or "COS" not in ai_data or "Asite" not in ai_data:
            logger.error("Invalid AI data format: missing 'COS' or 'Asite' keys")
            st.error("❌ Invalid AI data format for Excel generation.")
            return None

        logger.info(f"AI data keys: {list(ai_data.keys())}")
        logger.info(f"COS data: {ai_data.get('COS', [])}")
        logger.info(f"Asite data: {ai_data.get('Asite', [])}")

        # Normalize slab_data structure
        if isinstance(ai_data, dict):
            if "Slab" in ai_data:
                slab_counts = ai_data["Slab"]
                logger.info("Slab data found with 'Slab' key.")
            else:
                slab_counts = ai_data
                logger.info("Slab data provided directly without 'Slab' key.")
        else:
            logger.warning("Invalid slab data format. Expected a dictionary. Proceeding without slab data.")
            slab_counts = {}
        
        logger.info(f"Slab counts: {slab_counts}")

        # Define the COS to Asite activity name mapping
        cos_to_asite_mapping = {
            "EL-First Fix": "Wall Conducting",
            "Installation of doors": ["Door/Window Frame", "Door/Window Shutter"],
            "Min. count of UP-First Fix and CP-First Fix": "Plumbing Works",  
            "Water Proofing Works": "Waterproofing - Sunken",
            "Gypsum & POP Punning": "POP & Gypsum Plaster",
            "Wall Tile": "Wall Tile",
            "Floor Tile": "Floor Tile",
            "EL-Second Fix": "Wiring & Switch Socket",
            "Sewer Line": "Sewer Line",
            "Line Storm Line": "Rain Water/Storm",
            "GSB": "Granular Sub-base",
            "WMM": "WMM",
            "Saucer drain": "Saucer drain/Paver block",
            "Kerb Stone": "Kerb Stone",
            "Electrical": "Electrical Cable",
            "Concreting": "Concreting"
        }

        # Activities that map to "Concreting" in Asite
        slab_cast_activities = ["Shuttering", "Reinforcement", "Concreting"]  

        # Initialize lists to store data
        consolidated_rows = []

        # Process Slab data
        slab_data_dict = {}
        for tower_name, total_count in slab_counts.items():
            if tower_name != "Tower Name" and tower_name != "Total":  # Skip header and total rows
                original_tower_name = tower_name  # Keep original for debugging
                
                # Normalize tower name but preserve A/B suffixes
                if "Tower" in tower_name:
                    tower_name = tower_name.replace("Tower ", "T").replace("(", "").replace(")", "")
                
                logger.info(f"Processing Slab Tower: {original_tower_name} -> {tower_name}")
                
                count = int(total_count) if pd.notna(total_count) else 0
                
                # Special handling for Tower 4 - split slab data between T4A and T4B
                if tower_name == "T4":
                    half_count = count // 2
                    remainder = count % 2
                    
                    slab_data_dict["T4A"] = half_count + remainder
                    slab_data_dict["T4B"] = half_count
                    
                    logger.info(f"Split T4 Slab data: T4A={half_count + remainder}, T4B={half_count}")
                else:
                    slab_data_dict[tower_name] = count

        logger.info(f"Processed slab_data_dict: {slab_data_dict}")

        # Process COS data
        cos_data_dict = {}
        for tower_data in ai_data.get("COS", []):
            tower_name = tower_data.get("Tower", "Unknown Tower")
            original_tower_name = tower_name  # Keep original for debugging
            
            if "Tower" in tower_name:
                tower_name = tower_name.replace("Tower ", "T").replace("(", "").replace(")", "")
            
            logger.info(f"Processing COS Tower: {original_tower_name} -> {tower_name}")
            
            for category_data in tower_data.get("Categories", []):
                category = category_data.get("Category", "Unknown Category")
                if category == "ED Civil":
                    category = "Civil Works"
                elif category == "MEP":
                    category = "MEP Works"
                elif category == "Interior Finishing":
                    category = "Interior Finishing Works"
                elif category == "Structure Work":
                    category = "Structure Works"
                for activity in category_data.get("Activities", []):
                    activity_name = activity.get("Activity Name", "Unknown Activity")
                    count = int(activity.get("Total", 0)) if pd.notna(activity.get("Total")) else 0
                    open_missing = activity.get("OpenMissingOverride", None)
                    
                    if tower_name == "T4":
                        half_count = count // 2
                        remainder = count % 2
                        
                        key_4a = ("T4A", activity_name, category)
                        cos_data_dict[key_4a] = {"count": half_count + remainder, "open_missing": open_missing}
                        
                        key_4b = ("T4B", activity_name, category)
                        cos_data_dict[key_4b] = {"count": half_count, "open_missing": open_missing}
                        
                        logger.info(f"Split T4 {activity_name}: T4A={half_count + remainder}, T4B={half_count}")
                    else:
                        key = (tower_name, activity_name, category)
                        cos_data_dict[key] = {"count": count, "open_missing": open_missing}

        logger.info(f"COS data dict: {cos_data_dict}")

        # Process Asite data
        asite_data_dict = {}
        for tower_data in ai_data.get("Asite", []):
            tower_name = tower_data.get("Tower", "Unknown Tower")
            original_tower_name = tower_name  # Keep original for debugging
            
            if "Tower" in tower_name:
                tower_name = tower_name.replace("Tower ", "T").replace("(", "").replace(")", "")
            
            logger.info(f"Processing Asite Tower: {original_tower_name} -> {tower_name}")
            
            for category_data in tower_data.get("Categories", []):
                category = category_data.get("Category", "Unknown Category")
                if category == "ED Civil":
                    category = "Civil Works"
                elif category == "MEP":
                    category = "MEP Works"
                elif category == "Interior Finishing":
                    category = "Interior Finishing Works"
                elif category == "Structure Work":
                    category = "Structure Works"
                for activity in category_data.get("Activities", []):
                    activity_name = activity.get("Activity Name", "Unknown Activity")
                    count = int(activity.get("Total", 0)) if pd.notna(activity.get("Total")) else 0
                    open_missing = activity.get("OpenMissingOverride", None)
                    
                    if tower_name == "T4":
                        half_count = count // 2
                        remainder = count % 2
                        
                        key_4a = ("T4A", activity_name, category)
                        asite_data_dict[key_4a] = {"count": half_count + remainder, "open_missing": open_missing}
                        
                        key_4b = ("T4B", activity_name, category)
                        asite_data_dict[key_4b] = {"count": half_count, "open_missing": open_missing}
                        
                        logger.info(f"Split T4 Asite {activity_name}: T4A={half_count + remainder}, T4B={half_count}")
                    else:
                        key = (tower_name, activity_name, category)
                        asite_data_dict[key] = {"count": count, "open_missing": open_missing}

        logger.info(f"Asite data dict: {asite_data_dict}")

        # Normalize COS data to use Asite activity names
        normalized_cos_data = {}
        for (tower, cos_activity, category), data in cos_data_dict.items():
            count = data["count"]
            open_missing = data["open_missing"]
            
            # Map slab-related activities to "Concreting"
            if cos_activity in slab_cast_activities or cos_activity == "Slab Conducting":
                asite_activity = "Concreting"
                key = (tower, asite_activity, category)
                existing_data = normalized_cos_data.get(key, {"count": 0, "open_missing": None})
                normalized_cos_data[key] = {
                    "count": existing_data["count"] + count,
                    "open_missing": open_missing if open_missing is not None else existing_data["open_missing"]
                }
                logger.info(f"Mapped {cos_activity} to Concreting for {tower} in {category}")
            elif cos_activity in ["UP-First Fix", "CP-First Fix"]:
                asite_activity = "Plumbing Works"
                key = (tower, asite_activity, category)
                existing_data = normalized_cos_data.get(key, {"count": float('inf'), "open_missing": None})
                normalized_cos_data[key] = {
                    "count": min(existing_data["count"], count) if existing_data["count"] != float('inf') else count,
                    "open_missing": open_missing if open_missing is not None else existing_data["open_missing"]
                }
            elif cos_activity in cos_to_asite_mapping and isinstance(cos_to_asite_mapping[cos_activity], list):
                for asite_activity in cos_to_asite_mapping[cos_activity]:
                    key = (tower, asite_activity, category)
                    normalized_cos_data[key] = {"count": count, "open_missing": open_missing}
            elif cos_activity in cos_to_asite_mapping:
                asite_activity = cos_to_asite_mapping[cos_activity]
                key = (tower, asite_activity, category)
                normalized_cos_data[key] = {"count": count, "open_missing": open_missing}
            else:
                logger.warning(f"No Asite mapping found for COS activity: {cos_activity}")
                key = (tower, cos_activity, category)
                normalized_cos_data[key] = {"count": count, "open_missing": open_missing}

        logger.info(f"Normalized COS data before slab merge: {normalized_cos_data}")

        # **CRITICAL UPDATE**: Merge slab data with normalized COS data for "Concreting" activities
        # This ensures slab_data_dict counts are properly mapped to "Concreting" in Completed Work column
        
        # First, identify all categories that have "Concreting" activities from Asite data
        concreting_categories = {}
        for (tower, activity, category) in asite_data_dict.keys():
            if activity == "Concreting":
                if tower not in concreting_categories:
                    concreting_categories[tower] = []
                if category not in concreting_categories[tower]:
                    concreting_categories[tower].append(category)
        
        logger.info(f"Found Concreting categories: {concreting_categories}")
        
        # Update/Create Concreting entries with slab data for each tower and relevant category
        for tower_name, slab_count in slab_data_dict.items():
            # Get categories for this tower, default to common categories if not found
            categories = concreting_categories.get(tower_name, ["MEP Works", "Structure Works"])
            
            for category in categories:
                key = (tower_name, "Concreting", category)
                
                # FORCE UPDATE: Set the slab count as the completed work count for Concreting
                # This will override any existing count for Concreting with the actual slab count
                normalized_cos_data[key] = {
                    "count": slab_count,  # This will show in "Completed Work*(Count of Flat)" column
                    "open_missing": normalized_cos_data.get(key, {}).get("open_missing", None)
                }
                logger.info(f"UPDATED: Set Concreting for {tower_name} in {category} with slab count: {slab_count}")

        # Also ensure any tower in slab_data_dict that doesn't have existing Concreting entries gets them
        for tower_name, slab_count in slab_data_dict.items():
            # Check if this tower has any Concreting entries
            has_concreting = any(key[0] == tower_name and key[1] == "Concreting" for key in normalized_cos_data.keys())
            
            if not has_concreting:
                # Create Concreting entry with default category
                default_category = "Structure Works"  # or determine based on your business logic
                key = (tower_name, "Concreting", default_category)
                normalized_cos_data[key] = {
                    "count": slab_count,
                    "open_missing": None
                }
                logger.info(f"CREATED: New Concreting entry for {tower_name} in {default_category} with slab count: {slab_count}")

        logger.info(f"After merging slab data, normalized COS data: {normalized_cos_data}")

        # Combine normalized COS (including slab) and Asite data
        all_keys = set(normalized_cos_data.keys()).union(set(asite_data_dict.keys()))
        for key in all_keys:
            tower_name, activity_name, category = key
            # Skip "No. of Slab cast" to avoid displaying it (since we're now using "Concreting")
            if activity_name == "No. of Slab cast":
                continue
                
            cos_data = normalized_cos_data.get(key, {"count": 0, "open_missing": None})
            asite_data = asite_data_dict.get(key, {"count": 0, "open_missing": None})
            cos_count = cos_data["count"]
            asite_count = asite_data["count"]
            
            # Special logging for Concreting activities to debug
            if activity_name == "Concreting":
                logger.info(f"CONCRETING DEBUG - Tower: {tower_name}, Category: {category}, COS Count: {cos_count}, Asite Count: {asite_count}")
            
            open_missing_override = cos_data["open_missing"] if cos_data["open_missing"] is not None else asite_data["open_missing"]
            if open_missing_override is not None:
                open_missing_count = open_missing_override
            else:
                open_missing_count = abs(cos_count - asite_count)
            in_progress_count = 0
            consolidated_rows.append({
                "Tower": tower_name,
                "Category": category,
                "Activity Name": activity_name,
                "Completed Work*(Count of Flat)": cos_count, 
                "In Progress Work*(Count of Flat)":in_progress_count,
                "Closed checklist against completed work": asite_count,
                "Open/Missing check list": open_missing_count
            })

        logger.info(f"Consolidated rows: {consolidated_rows}")

        # Create DataFrame
        df = pd.DataFrame(consolidated_rows)
        if df.empty:
            logger.warning("DataFrame is empty after processing. No data available to generate consolidated checklist.")
            st.warning("No data available to generate consolidated checklist.")
            # Create a minimal Excel file with a message
            output = BytesIO()
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Consolidated Checklist"
            worksheet.cell(row=1, column=1).value = "No data available to generate consolidated checklist."
            workbook.save(output)
            output.seek(0)
            return output

        logger.info(f"DataFrame created with {len(df)} rows")

        # Sort by Tower, Category, and Activity Name for consistency
        df.sort_values(by=["Tower", "Category", "Activity Name"], inplace=True)

        # Create a BytesIO buffer for the Excel file
        output = BytesIO()
        workbook = Workbook()

        # Remove default sheet created by openpyxl
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        # Define styles
        header_font = Font(bold=True)
        category_font = Font(bold=True, italic=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center')

        # Create Sheet 1: Consolidated Checklist
        worksheet1 = workbook.create_sheet(title="Consolidated Checklist")
        current_row = 1

        grouped_by_tower = df.groupby('Tower')

        for tower, tower_group in grouped_by_tower:
            worksheet1.cell(row=current_row, column=6).value = tower
            worksheet1.cell(row=current_row, column=6).font = header_font
            current_row += 1

            grouped_by_category = tower_group.groupby('Category')

            for category, cat_group in grouped_by_category:
                worksheet1.cell(row=current_row, column=6).value = f"{tower} June Checklist Status - {category}"
                worksheet1.cell(row=current_row, column=6).font = category_font
                current_row += 1

                headers = [
                    "ACTIVITY NAME",
                    "Completed Work*(Count of Flat)",
                    "In Progress Work*(Count of Flat)",
                    "Closed checklist against completed work",
                    "Open/Missing check list"
                ]
                for col, header in enumerate(headers, start=6):
                    cell = worksheet1.cell(row=current_row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = center_alignment

                current_row += 1

                for _, row in cat_group.iterrows():
                    worksheet1.cell(row=current_row, column=6).value = row["Activity Name"]
                    worksheet1.cell(row=current_row, column=7).value = row["Completed Work*(Count of Flat)"]
                    worksheet1.cell(row=current_row, column=8).value = row["In Progress Work*(Count of Flat)"]
                    worksheet1.cell(row=current_row, column=9).value = row["Closed checklist against completed work"]
                    worksheet1.cell(row=current_row, column=10).value = row["Open/Missing check list"]
                    for col in range(6, 10):
                        cell = worksheet1.cell(row=current_row, column=col)
                        cell.border = border
                        cell.alignment = center_alignment
                    current_row += 1

                total_open_missing = cat_group["Open/Missing check list"].sum()
                worksheet1.cell(row=current_row, column=6).value = "TOTAL pending checklist June"
                worksheet1.cell(row=current_row, column=9).value = total_open_missing
                for col in range(6, 10):
                    cell = worksheet1.cell(row=current_row, column=col)
                    cell.font = category_font
                    cell.border = border
                    cell.alignment = center_alignment
                current_row += 1

            current_row += 1

        # Adjust column widths for Sheet 1
        for col in worksheet1.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet1.column_dimensions[column].width = adjusted_width

        # Create Sheet 2: Summary Checklist
        worksheet2 = workbook.create_sheet(title="Checklist June")
        current_row = 1

        # Write title
        worksheet2.cell(row=current_row, column=1).value = "Checklist: June"
        worksheet2.cell(row=current_row, column=1).font = header_font
        current_row += 1

        # Write headers
        headers = [
            "Site",
            "Total of Missing & Open Checklist-Civil",
            "Total of Missing & Open Checklist-MEP",
            "TOTAL"
        ]
        for col, header in enumerate(headers, start=1):
            cell = worksheet2.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment
        current_row += 1

        # Categorize towers into Civil and MEP
        def map_category_to_type(category):
            if category in ["Civil Works", "Structure Works"]:
                return "Civil"
            elif category in ["MEP Works", "Interior Finishing Works"]:
                return "MEP"
            elif category == "External Development":
                return "Civil"  # Assuming External Development is Civil
            else:
                return "Civil"  # Default to Civil if unknown

        # Aggregate open/missing counts by tower and type (Civil/MEP)
        summary_data = {}
        for _, row in df.iterrows():
            tower = row["Tower"]
            category = row["Category"]
            open_missing = row["Open/Missing check list"]
            
            # Convert tower name to display format (e.g., "T4A" -> "Veridia-Tower 04 A")
            if "External Development" in category:
                site_name = f"External Development-{tower}"
            else:
                tower_num = tower[1:]  # Remove 'T' prefix
                if len(tower_num) == 1:
                    tower_num = f"0{tower_num}"  # Pad single digits (e.g., "T2" -> "02")
                site_name = f"Veridia-Tower {tower_num}"

            type_ = map_category_to_type(category)
            
            if site_name not in summary_data:
                summary_data[site_name] = {"Civil": 0, "MEP": 0}
            
            summary_data[site_name][type_] += open_missing

        logger.info(f"Summary data for Sheet 2: {summary_data}")

        # Write summary data to Sheet 2
        for site_name, counts in sorted(summary_data.items()):
            civil_count = counts["Civil"]
            mep_count = counts["MEP"]
            total_count = civil_count + mep_count
            
            worksheet2.cell(row=current_row, column=1).value = site_name
            worksheet2.cell(row=current_row, column=2).value = civil_count
            worksheet2.cell(row=current_row, column=3).value = mep_count
            worksheet2.cell(row=current_row, column=4).value = total_count
            
            for col in range(1, 5):
                cell = worksheet2.cell(row=current_row, column=col)
                cell.border = border
                cell.alignment = center_alignment
            current_row += 1

        # Adjust column widths for Sheet 2
        for col in worksheet2.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet2.column_dimensions[column].width = adjusted_width

        # Save the workbook to the BytesIO buffer
        workbook.save(output)
        output.seek(0)

        logger.info("Excel file generated successfully")
        return output

    except Exception as e:
        logger.error(f"Error generating consolidated Excel: {str(e)}", exc_info=True)
        st.error(f"❌ Error generating Excel file: {str(e)}")
        return None

# Streamlit UI - Modified Button Code
st.sidebar.title("🔒Asite Initialization")
email = st.sidebar.text_input("Email", "impwatson@gadieltechnologies.com", key="email_input")
password = st.sidebar.text_input("Password", "Srihari@790$", type="password", key="password_input")

if st.sidebar.button("Initialize and Fetch Data"):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        success = loop.run_until_complete(initialize_and_fetch_data(email, password))
        if success:
            st.sidebar.success("Initialization and data fetching completed successfully!")
        else:
            st.sidebar.error("Initialization and data fetching failed!")
    except Exception as e:
        st.sidebar.error(f"Initialization and data fetching failed: {str(e)}")
    finally:
        loop.close()

# Combined function to handle both analysis and activity count display
@function_timer(show_args=True)
def run_analysis_and_display():
    try:
        st.write("Running status analysis...")
        AnalyzeStatusManually()
        st.success("Status analysis completed successfully!")

        st.write("Processing AI data totals...")
        if 'ai_response' not in st.session_state or not st.session_state.ai_response:
            st.error("❌ No AI data available to process totals. Please ensure analysis ran successfully.")
            return

        st.write("Displaying activity counts...")
        display_activity_count()
        st.success("Activity counts displayed successfully!")

        st.write("Generating consolidated checklist Excel file...")
        # Defensive initialization for ai_response
        if 'ai_response' not in st.session_state or not st.session_state.ai_response:
            st.session_state.ai_response = {"COS": [], "Asite": []}
            st.warning("AI response not found. Initialized to empty data.")

        # Ensure GetSlabReport is called to populate slabreport
        GetSlabReport()

        # Defensive initialization for slabreport
        if 'slabreport' not in st.session_state:
            st.session_state.slabreport = {}
            st.warning("Slab report data not found. Initialized to empty dictionary.")

        with st.spinner("Generating Excel file... This may take a moment."):
            excel_file = generate_consolidated_Checklist_excel(st.session_state.ai_response)
        
        if excel_file:
            timestamp = pd.Timestamp.now(tz='Asia/Kolkata').strftime('%Y%m%d_%H%M')
            file_name = f"Consolidated_Checklist_Veridia_{timestamp}.xlsx"
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.sidebar.download_button(
                    label="📥 Download Checklist Excel",
                    data=excel_file,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_button",
                    help="Click to download the consolidated checklist in Excel format."
                )
            st.success("Excel file generated successfully! Click the button above to download.")
        else:
            st.error("Failed to generate Excel file. Please check the logs for details.")

    except Exception as e:
        st.error(f"Error during analysis, display, or Excel generation: {str(e)}")
        logging.error(f"Error during analysis, display, or Excel generation: {str(e)}")

st.sidebar.title("📊 Status Analysis")

if st.sidebar.button("Analyze and Display Activity Counts"):
    run_analysis_and_display()

st.sidebar.title("📊 Slab Cycle")
st.session_state.ignore_year = st.sidebar.number_input("Ignore Year", min_value=1900, max_value=2100, value=2023, step=1, key="ignore_year1")
st.session_state.ignore_month = st.sidebar.number_input("Ignore Month", min_value=1, max_value=12, value=3, step=1, key="ignore_month1")

